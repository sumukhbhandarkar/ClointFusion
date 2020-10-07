# Project Name: ClointFusion
# Project Description: A Python based RPA Automation Framework for Desktop GUI, Citrix, Web and basic Excel operations.
import subprocess
import os
import time
import sys
import platform
import urllib.request

current_environment = 0


os_name = str(platform.system()).lower()
c_drive_base_dir = ""

if os_name == 'windows':
    c_drive_base_dir = r"C:\ClointFusion\My_Bot"  

current_working_dir = os.path.dirname(os.path.realpath(__file__)) #get cwd
cf_icon_file_path = os.path.join((current_working_dir),"CF_ICON.ico")

bot_name = ""
enable_semi_automatic_mode = False


if os_name == 'windows':
    os_path=os.environ['USERPROFILE']
    win_venv_scripts_folder_path = (r"{}\ClointFusion\Scripts".format(os_path))
    
    win_venv_python_path = os.path.join(win_venv_scripts_folder_path, "python.exe")
    
    env_pip_path = os.path.join(win_venv_scripts_folder_path,"pip")

    if os.path.exists(r"{}\ClointFusion\cf_venv_activated.txt".format(os_path)) == False:
        subprocess.call(r"cmd.exe -ArgumentList /c python -m venv {}\ClointFusion & {}\ClointFusion\Scripts\activate & python.exe -m pip install -U pip & python -m pip install -U wheel & python -m pip install -U ClointFusion & deactivate & type nul > {}\ClointFusion\cf_venv_activated.txt".format(os_path,os_path,os_path))

        while True:
            if os.path.exists(r"{}\ClointFusion\cf_venv_activated.txt".format(os_path)):
                break

    if os.path.exists(r"{}\ClointFusion\Scripts\activate_this.py".format(os_path)) == False :
        with open(r"{}\ClointFusion\Scripts\activate_this.py".format(os_path), 'w') as f:
            activate_this_py =""" 
    try:
        __file__
    except NameError:
        raise AssertionError(
            "You must run this like execfile('path/to/activate_this.py', dict(__file__='path/to/activate_this.py'))")
    import sys
    import os

    old_os_path = os.environ.get('PATH', '')
    os.environ['PATH'] = os.path.dirname(os.path.abspath(__file__)) + os.pathsep + old_os_path
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if sys.platform == 'win32':
        site_packages = os.path.join(base, 'Lib', 'site-packages')
    else:
        site_packages = os.path.join(base, 'lib', 'python%s' % sys.version[:3], 'site-packages')
    prev_sys_path = list(sys.path)
    import site
    site.addsitedir(site_packages)
    sys.real_prefix = sys.prefix
    sys.prefix = base
    # Move the added items to the front of the path:
    new_sys_path = []
    for item in list(sys.path):
        if item not in prev_sys_path:
            new_sys_path.append(item)
            sys.path.remove(item)
    sys.path[:0] = new_sys_path """
            f.write(activate_this_py)

    subprocess.call(r"cmd.exe -ArgumentList /c {}\ClointFusion\Scripts\activate".format(os_path))

    activate_venv = r"{}\ClointFusion\Scripts\activate_this.py".format(os_path)
    exec(open(activate_venv).read(), {'__file__': activate_venv})

list_of_required_packages = ["wheel","urllib3","beautifulsoup4","pdfplumber","watchdog","wordcloud","scipy","numpy","howdoi","seaborn","texthero","emoji","helium","kaleido", "folium", "zipcodes", "plotly", "PyAutoGUI", "PyGetWindow", "XlsxWriter" ,"PySimpleGUI", "chromedriver-autoinstaller", "imutils", "keyboard", "joblib", "opencv-python", "python-imageseach-drov0", "openpyxl", "pandas", "pif", "pytesseract", "scikit-image", "selenium", "xlrd", "clipboard"]

#decorator to push a function to background using asyncio
def background(f):
    """
    Decorator function to push a function to background using asyncio
    """
    import asyncio
    try:
        from functools import wraps
        @wraps(f)
        def wrapped(*args, **kwargs):
            loop = asyncio.get_event_loop()
            if callable(f):
                return loop.run_in_executor(None, f, *args, **kwargs)
            else:
                raise TypeError('Task must be a callable')    
        return wrapped
    except Exception as ex:
        print("Task pushed to background = "+str(f) + str(ex))

@background
def _download_cloint_ico():    
    """
    Internal function to download ClointFusion ICON from GitHub
    """
    if not os.path.exists(cf_icon_file_path):
        urllib.request.urlretrieve('https://raw.githubusercontent.com/ClointFusion/ClointFusion/master/Cloint-ICON.ico',cf_icon_file_path)

import emoji
from pandas.core.algorithms import mode
from xlrd.formula import colname

def show_emoji(strInput=""):
    """
    Function which prints Emojis

    Usage: 
    print(show_emoji('thumbsup'))
    print("OK",show_emoji('thumbsup'))
    Default: thumbsup
    """
    if not strInput:
        return(emoji.emojize(":{}:".format(str('thumbsup').lower()),use_aliases=True,variant="emoji_type"))
    else:
        return(emoji.emojize(":{}:".format(str(strInput).lower()),use_aliases=True,variant="emoji_type"))

@background
def load_missing_python_packages():
    """
    Installs missing python packages
    """       

    subprocess.call("powershell Set-ExecutionPolicy -ExecutionPolicy RemoteSigned -Scope CurrentUser")
    
        
    try:
        import PySimpleGUI
    except:
        os.system("{} install --upgrade {}".format(sys.executable,'PySimpleGUI'))

    #install missing packages
    try:
        reqs = subprocess.check_output([win_venv_python_path, '-m', 'pip', 'list'])
        installed_packages = [r.decode().split('==')[0] for r in reqs.split()]

        missing_packages = ' '.join(list(set(list_of_required_packages)-set(installed_packages)))

        if missing_packages:
            print("{} package(s) are missing".format(missing_packages)) 
            
            os.system("{} -m pip install --upgrade pip".format(sys.executable))
            
            os.system("{} install --upgrade {}".format(env_pip_path,missing_packages)) 


    except Exception as ex:
        print("Error in load_missing_python_packages="+str(ex))

def is_execution_required_today(file_name,save_todays_date=False):
    """
    Function which ensutres that a another function which calls this function is executed only once per day.
    Returns boolean True/False if another function to be executed today or not
    """
    from datetime import datetime
    last_updated_date_file = r"{}\ClointFusion\Scripts\{}.txt".format(os_path,file_name)
    EXECUTE_TODAY = False
    last_updated_on_date = ""

    try:    
        with open(last_updated_date_file, 'r') as f:
            last_updated_on_date = str(f.read())
            save_todays_date = False
    except:
        save_todays_date = True

    if save_todays_date:
        with open(last_updated_date_file, 'w') as f:
            last_updated_on_date = datetime.today().strftime('%d')
            f.write(str(last_updated_on_date))
            EXECUTE_TODAY = True

    today_date = str(datetime.today().strftime('%d'))
    if last_updated_on_date != today_date:
        EXECUTE_TODAY = True

    return EXECUTE_TODAY

#upgrade existing packages
@background
def update_all_packages_in_cloint_fusion_virtual_environment():
    """
    Function to UPGRADE all packages related to ClointFusion. This function runs in background and is silent.
    """
    EXECUTE_TODAY = is_execution_required_today('update_all_packages_in_cloint_fusion_virtual_environment')

    if EXECUTE_TODAY:
        try:
            updating_required_packages= ' '.join(list(set(list_of_required_packages)))
            
            _ = subprocess.run("{} install --upgrade {}".format(env_pip_path,updating_required_packages),capture_output=True)
            is_execution_required_today('update_all_packages_in_cloint_fusion_virtual_environment',True)

        except Exception as ex:
            print("Error in update_all_packages_cloint_fusion_virtual_environment="+str(ex))

def _welcome_to_clointfusion():
    """
    Internal Function to display welcome message & push a notification to ClointFusion Slack
    """
    welcome_msg = "Welcome to ClointFusion, Made in India with " + show_emoji('red_heart')
    print(welcome_msg)


def string_remove_special_characters(inputStr=""):
    """
    Removes all the special character.

    Parameters:
        inputStr  (str) : string for removing all the special character in it.

    Returns :
        outputStr (str) : returns the alphanumeric string.
    """

    if inputStr:
        outputStr = ''.join(e for e in inputStr if e.isalnum())
        return outputStr  

def _set_bot_name(strBotName=""):
    """
    Internal function
    If a botname is given, it will be used in the log file and in Task Scheduler
    we can also access the botname variable globally.

    Parameters :
        strBotName (str) : Name of the bot
    """
    global c_drive_base_dir
    global bot_name

    if not strBotName: #if user has not given bot_name
        bot_name = current_working_dir[current_working_dir.rindex("\\") + 1 : ] #Assumption that user has given proper folder name and so taking it as BOT name

    else:
        strBotName = string_remove_special_characters(strBotName)    
        bot_name = strBotName

    c_drive_base_dir = c_drive_base_dir + "_" + bot_name
    
def folder_create(strFolderPath=""):
    """
    while making leaf directory if any intermediate-level directory is missing,
    folder_create() method will create them all.

    Parameters:
        folderPath (str) : path to the folder where the folder is to be created.

    For example consider the following path:
        
        Suppose we want to create directory ‘krishna’ but Directory ‘project’ and ‘python’ are unavailable in the path.
        Then folder_create() method will create all unavailable/missing directory in the specified path.
        ‘project’ and ‘python’ will be created first then ‘krishna’ directory will be created.
    """
    
    try:
        if not strFolderPath:
            strFolderPath = gui_get_any_input_from_user('folder path to Create folder')

        if not os.path.exists(strFolderPath):
            os.makedirs(strFolderPath)
    except Exception as ex:
        print("Error in folder_create="+str(ex))

def _create_status_log_file(xtLogFilePath):
    """
    Internal Function to create Status Log File
    """
    if not os.path.exists(xtLogFilePath):
        df = pd.DataFrame({'Timestamp': [], 'Status':[]})
        writer = pd.ExcelWriter(xtLogFilePath, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        writer.close()

# @timeit
def _init_log_file():
    """
    Generates the log and saves it to the file in the given base directory. Internal function
    """
    global log_path
    global fullPathToStatusLogFile
    
    try:
        if bot_name:
            excelFileName = str(bot_name) + "-StatusLog.xlsx"
        else:
            excelFileName = "StatusLog.xlsx"

        fullPathToStatusLogFile = os.path.join(folderPathToStatusLogFile,excelFileName)
        
        folder_create(folderPathToStatusLogFile)
        
        _create_status_log_file(fullPathToStatusLogFile)   

    except Exception as ex:
        print("ERROR in _init_log_file="+str(ex))

def folder_read_text_file(txt_file_path=""):
    """
    Reads from a given text file and returns entire contents as a single list
    """
    try:
        if not txt_file_path:
            txt_file_path = gui_get_any_file_from_user('the text file to READ from',"txt")

        with open(txt_file_path) as f:
            file_contents = f.readlines()
        return file_contents
    except:
        return None

def folder_write_text_file(txt_file_path="",contents=""):
    """
    Writes given contents to a text file
    """
    try:
        
        if not txt_file_path:
            txt_file_path = gui_get_any_file_from_user('the text file to WRITE to',"txt")

        if not contents:
            contents = gui_get_any_input_from_user('text file contents')

        f = open(txt_file_path,'w')
        f.writelines(str(contents))
        f.close()
        
    except Exception as ex:
        print("Error in folder_write_text_file="+str(ex))

def _ask_user_semi_automatic_mode():
    """
    Ask user to 'Enable Semi Automatic Mode'
    """
    global enable_semi_automatic_mode
    values = []
    
    stored_do_not_ask_user_preference = folder_read_text_file(os.path.join(config_folder_path, 'Dont_Ask_Again.txt'))
    enable_semi_automatic_mode = folder_read_text_file(os.path.join(config_folder_path, 'Semi_Automatic_Mode.txt'))

    if enable_semi_automatic_mode:
        enable_semi_automatic_mode = enable_semi_automatic_mode[0]
    
    bot_config_path = os.path.join(config_folder_path,bot_name + ".xlsx")

    if stored_do_not_ask_user_preference is None or str(stored_do_not_ask_user_preference[0]).lower() == 'false':

        layout = [[sg.Text('Do you want me to store GUI responses & use them next time when you run this BOT ?',text_color='orange',font='Courier 13')],
                [sg.Submit('Yes',bind_return_key=True,button_color=('white','green'),font='Courier 14'), sg.CloseButton('No', button_color=('white','firebrick'),font='Courier 14')],
                [sg.Checkbox('Do not ask me again', key='-DONT_ASK_AGAIN-',default=False, text_color='yellow',enable_events=True)],
                [sg.Text("To see this message again, goto 'Config_Files' folder of your BOT and change 'Dont_Ask_Again.txt' to False. \n Please find path here: {}".format(os.path.join(config_folder_path, 'Dont_Ask_Again.txt')),key='-DND-',visible=False,font='Courier 8')]]

        window = sg.Window('ClointFusion - Enable Semi Automatic Mode ?',layout,return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)
        folder_write_text_file(os.path.join(config_folder_path, 'Dont_Ask_Again.txt'),str(False))

        while True:
            event, values = window.read()
            if event == '-DONT_ASK_AGAIN-':
                stored_do_not_ask_user_preference = values['-DONT_ASK_AGAIN-']
                folder_write_text_file(os.path.join(config_folder_path, 'Dont_Ask_Again.txt'),str(stored_do_not_ask_user_preference))

                if values['-DONT_ASK_AGAIN-']:
                    window.Element('-DND-').Update(visible=True)
                else:
                    window.Element('-DND-').Update(visible=False)
                    
            if event in (sg.WIN_CLOSED, 'No'): #ask me every time
                enable_semi_automatic_mode = False
                break
            elif event == 'Yes': #do not ask me again
                enable_semi_automatic_mode = True
                break
    
        window.close()

        if not os.path.exists(bot_config_path):
            df = pd.DataFrame({'SNO': [],'KEY': [], 'VALUE':[]})
            writer = pd.ExcelWriter(bot_config_path, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            writer.close()

        if enable_semi_automatic_mode:
            print("Semi Automatic Mode is ENABLED "+ show_emoji())
        else:
            print("Semi Automatic Mode is DISABLED "+ show_emoji())
        
        folder_write_text_file(os.path.join(config_folder_path, 'Semi_Automatic_Mode.txt'),str(enable_semi_automatic_mode))

_download_cloint_ico()

load_missing_python_packages()
update_all_packages_in_cloint_fusion_virtual_environment()

from unicodedata import name
import pyautogui as pg
import pygetwindow as gw
import json
import time
import pandas as pd
import keyboard as kb
import PySimpleGUI as sg
import os
import xlrd
import openpyxl as op
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import subprocess

from functools import lru_cache
import threading
from threading import Timer
import traceback

import socket
from cv2 import cv2
import base64
import pytesseract
import signal
from skimage.measure import compare_ssim
from skimage.metrics import structural_similarity
import imutils
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
import chromedriver_autoinstaller
from selenium.webdriver.support.expected_conditions import _find_elements
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from python_imagesearch.imagesearch import imagesearch
from python_imagesearch.imagesearch import region_grabber
from python_imagesearch.imagesearch import imagesearcharea
import shutil
from shutil import which
from joblib import Parallel, delayed
from concurrent.futures import ThreadPoolExecutor, as_completed
import clipboard
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

import plotly.express as px
from kaleido.scopes.plotly import PlotlyScope
import plotly.graph_objects as go
import zipcodes
import folium
from json import (load as jsonload, dump as jsondump)
from helium import *
from os import link
from selenium.webdriver import ChromeOptions
import dis
import texthero as hero
from texthero import preprocessing
import seaborn as sb
import matplotlib.pyplot as plt
from urllib.request import urlopen 
from hashlib import sha256
from PIL import Image
import numpy as np
from scipy.ndimage import gaussian_gradient_magnitude
from wordcloud import WordCloud, ImageColorGenerator
from bs4 import BeautifulSoup
import requests
import pdfplumber

import math
import watchdog.events
import watchdog.observers
from PyQt5 import QtWidgets, QtCore, QtGui
import tkinter as tk
from PIL import ImageGrab
sg.theme('Dark') 

# ########################
# ClointFusion's DEFAULT SERVICES

_set_bot_name()
_welcome_to_clointfusion()
folder_create(c_drive_base_dir) 

log_path = os.path.join(c_drive_base_dir, "Logs")
img_folder_path =  os.path.join(c_drive_base_dir, "Images") 
batch_file_path = os.path.join(c_drive_base_dir, "Batch_File") 
config_folder_path = os.path.join(c_drive_base_dir, "Config_Files") 
output_folder_path = os.path.join(c_drive_base_dir, "Output") 
error_screen_shots_path = os.path.join(c_drive_base_dir, "Error_Screenshots")
folderPathToStatusLogFile = os.path.join(c_drive_base_dir,"StatusLogExcel")
Cloint_PNG_Logo_Path = os.path.join(img_folder_path,"Cloint_Logo.PNG")
client_secret_gmail_json_path = os.path.join(config_folder_path, "cs.json")

folder_create(log_path)
folder_create(img_folder_path)
folder_create(batch_file_path)
folder_create(config_folder_path)
folder_create(error_screen_shots_path)
folder_create(output_folder_path)
_init_log_file()

_ask_user_semi_automatic_mode()

# ########################

# Global VARIABLES
# #Web Browser Automation Global Variables
service = ""
driver = ""
URL = ""
Chrome_Service_Started = False
cloint_logo_base64 = b'iVBORw0KGgoAAAANSUhEUgAAAMUAAADHCAYAAACp8Jf7AAAABGdBTUEAALGPC/xhBQAAAAlwSFlzAAAOxAAADsQBlSsOGwAAONRJREFUeF7tfYeLFGnX7/wD9/Ld9+O7+33vfffdXfOoYxzzuq5pV3ZFUZRFcXFdjJgwYcSMOWAOmCPmgGPEhIqKioqKioqKoigO03TTTTfV9Lnnd6pq7Ol5ZqZDdVdVT//gh+NMd3V11fnVOed5znOePMrBWkTCROEQUShAkUAJRUreU+Tjcwq/uUfhZ1dIe3CatDtHSbuxj0JXd1Ho0hYKXdz8lZe2knZtF2k3D5B29xhpj85S+Pk1Cr99QJFPLyji+UAU8Mjx5XPweTlYipwokkaEDdNPEe8n3ehf3BAjDl3YQMEjsyiwfRj51/Qg35IO5JvbirzTm5F3SiPyTmhI3nH1yTs2nzxj6pFndN1y9PLvvePyyTueXzeJXz+1MXlnNiffvNbkX9aJAhv6UGDXaAoem0+hyyyi+6co/Oo2RT6/oojvC5EWNM4xh2SQE0W84CdyxF9MkfePSXt4Vp7wgYMzyL+hP/kWtWejb0oeNuKSUbXJM7wGeYb+EEX8nzksQeI4oOpv4IhaLKI65J3YgEVTyILpSIGtg1gsc9kL7aTw00siWPEsEHEOcSEnigoRoQg8wYen4gGCJxaTf8vfugCmNJEnegkbesng73Wy8QvZWEGlESdCUxDxUN5jCBD/sljE08xsRr4V7Fl2DqfQ2VUs5jPsTV7mPEkVyIkiGmFNwo/wy1scBm2kwLah5FvQljwTC6hkZG0qGcJGb4oAP0MUJtkwo1nOyBOlyvgTIcRhckQNDtc4LJvWhHzLfqbAnrEU4rwFeYp4kVxeUgY5USAs4uQVSXDw9Eryr+3NxtOMw6A6urHD+E1GiyCWbMiOEgXIYiil/I4FAo6syZ6kHvnmtJBwK3RpM4Vfc06CfCQnkOorigg/IcMvb3JYsY7zgn7kgRBGGN5AQiHD2OMlG7KjRaGi8TrvmLrkm9uSAjuGUujKdt2DcOhYXVG9RAGv8OUthW4cYAMYSd7ZrfmpybnB4Dg8QVVkQ3adKEzitQiz+F/vhPrkW/wjBQ9OIe3BKSLvJ75u1StJrx6iwJzB24cUPLOafMu7kmd8Q92Q/0Z+kKIYTLIhu1YU0UR4BYGM4mSdc5DAhr7iPSRBx7xINUBWiyIS8JL2/BoFD88m79zWnCxzngARmJ5hCBswqDLyRMmGnBWiiCYEgn/H5ZOfE/TQmRUcWt3TJw6zGFkpChHDo/MU2DeZvDNbsJHWMoTABquiysgTJRty1okimjje2DocWrWj4PH5FH51i0jLTnFklSgi/ATTnl6mwJ7x5J1RyMZaUw+RVEKIpsrIEyUbclaLwiS8x5ja5FvYViYJUb6CoexsQnaIgmPd8Ou7FDg4nTwQQ6lnqMQ7RFNl5ImSDblaiGKk8a+Iow75FrE4Ti+jyKfnWSMOd4sCo0kfn1OwaAXnDG0SF4NJlZEnSjbkaiMKk+bvOKzyr+jECfk2ipS8w43R749L4VpRRHyfKXR9L/lWdqOSUeawKhtkMlQZeaJkQ652ooihd1IDCmwZSNrDIqKgz7hT7oP7RMEuOvz8uswzSPkFxJCKIECVkSdKNuTqLgrz795ZzSl4ZKYUT7pxjsNVooiUfKDguXXkndOaDa+KEaVEqDLyRMmGXO1FYXIE5xsYqeKQSru5X6oH3AR3iAKJ9NMr5N/0F3lG5yeXN1RGlZEnSjbknCjK0zu1gIIHJlP43QPJAd0Ax4si4v1MwfOb9EQaBmyVd4imysgTJRtyThQK4j1japN/ZRfSbh9yRa7hXFFwLBp++5ACu8ayK25AJX+zGAazwYEqw06FKiNPlGzIOVFUQn6vd2ZTCp6YT5HiN8ZNdiacKQotKEssfUu7sIFx7hAtiJwo4mesQVtBlcHHS+Qa4+tRYMsAqVB26ryG40SBmv7gufXknV6oCyBWEDlRxE+VUadKlbEnytG1+IHXnrRbBx1Zou4oUaBbRWD/VPKMiwmXYqky7FSoMvJEyYacE0UcHFVTJ//sndWMgmdWSPMHJ8EhouD84dVd8m/8k0pG1KlcEKDKsFOhysgTJRtyThRx0BSFKQxM+B2YrJemOwT2iyISJu3xBXanndlAFfmDiirDToUqI0+UbMg5UcTBaFEYwvCMq0P+rQMp/PY+DEK3Cxthryi0AIVuH9GHW1XGXxFVhp0KVUaeKNmQc6KIg7GiMIWBYdu1v1H42WXWhb3CsE8UnGCFruwg74wW8XmHaKoMOxWqjDxRsiHnRBEHVaIQ8t+QgC/rQOGHp20dmbJHFBDEhU3kmdI0cUGAKsNOhSojT5RsyDlRxEGlIKLIr0E5unbvmFQy2IGMiyIS9FHwzBryTGyUnCBAlWGnQpWRJ0o25Jwo4qBKCGWov843vyVptw/a4jEyKgoUhokgJhQkLwhQZdipUGXkiZINOSeKOKgUgor2CSNzogh6KXgagkjBQ5hUGXYqVBl5omRDzokiDioFUBFNYRzKaCiVGVGEAhQ6zznExMZUMihFQYAqw06FKiNPlGzIX/lDVHNlk2wQYux8s4fXYgOpzTe9jvSk9YyuV7YDOX6Pv6OBsrzeeK/ZBlOIY1VBlVGnSpWhJ0Kl8VdCfo9vQSsjx8hMlW36RaGFKHRlF3knNbFGEKDKsFNhrIEnSiljB/VGy54RNfV2+pMKyDezUJoy+9f0lN60wf2TKXh8AYXOrNb3o7iyQ/ap0K7vlU7h6LGEPrah0yspeGweBfZNlNaW/tW/kW9hO2nJj/b8njEsHojGFCD+zUZRgPw+JN/hR2cyMlybXlFwLKjdOkzeaYV6yPQ3G5DKyBOlyrBTocrQqyJEYDZaHsFPdRaAd0Fb8m/oS8GDU2WfCu3eia/7RviL9W7fEY6PcWNNymRVDEv/zk9GxNNo5ub7LGUwsg/GnSMUOruGBTOB/Gt7yr4V3okslJEsEtMzZZMoDPqWd5QNbNKNNIoiIvs4+Oa21T0EBJENooAI8C88wZxW5N80kIKnlpJ29ziF39zX63gylRiyyLBTEoSH4rrg8XnkX9+HvLMKpT9sWjyGytATocLY4+boWuRf153C7x4aFyA9SJsowq/vkG9ZF8NDuFgUZmjEPyP2R6dB/85RFLq2W3oeZVQEVcEUycsb0kkcYZdvNgtkLOcoMGiz418qVBl6IlQZe9zk94+pTYEdgyny5bXxpa1HWkSBcMG/YYBucNGCcJMoJDxicmiErbn8m/+S+D/y/om+ekxCHwcD7X/QWf3NXQqdX8cepJf0hkWP2NL8Q2X0VVFl6IlQaewJEMeYUI+Ch6bqWwekAZaLAieKdpWe4bXLC8INojA8A0aDfIs7UPDoPAo/vUzkL8G307+k6xARj4Yu4sH9k8i3oA1/P/YeyXiOWCNPlCpDT5hY+92Qxb5G6ueshrWiYPeN/R48YxoYeYSbRGGIYWy+9JIKnd/AsesjDo2yrNM2ezn0gQ0WLSXf0g4cWtVLzGuoDD0RKo08QY7GvyyMOc31oVqLvbalotDundRXzA363hCEC0RhdAYRz7DiFw6Rdkr455g8IV1gsaMvU+jsavItaW94DjZalRCiqTL0RKgy8kQJUQhrkG/5zxIiWgnLRIEmA75lXQ0hRJMNz6migCBG1iEvh0nwDJEvbyQWr1Zg8Uc+PKHgyUXkm9+ajQ67u1YSVqkMPRGqjDxRloqCOaaWnniXvDW+UOqwRBRoQxPYOYZDEGORkNNFATFwyIQ2/ZIzIEyqbmKIBXprvbhOwX3j9YRcJQhQZeiJUGXkiTJaFBiRmlSfgqeXWrbeO3VR8JNGSjjG1I/KI6LJRugkUbAgPGPry/a/2uPzMjGWw1dgkhGTg/613fk6GXMdThaFIQzvnELS7p/EN9C/SApIWRTas6vkndUqJo9woCjEO9Qk7+xWFESoVPKBz96to0npRkRmz7E5i3dGU10MpjhUhp4IVUaeKGNFIawhK/ciH58Z3yF5pCSKSPE7vdkADFUpCBB/i2KsgSdDldFXRniH0flyrmjOnHUjSukChyN4+vpX/8rXz8g1VIaeCFVGnigrEIVnfB0Oh2el3IUweVGg0A9rI9DbtUIvAbJR2ikKFoR3alN+6i1iEVuXjFUfRCj8/jHnGhPIO7lh6t5CZeSJUikKJsKoWU1Ju3dUzjtZJC0KFGZ557Shkr8qEwTIhmmLKPizkUwv+JG0mwdd0cPU0fB9odCF9eTj2F1p7PFSZeSJUiWIUnIYte73lMKopESBZMy/fYRuoEohRBOviWKsgSdDpQiiiPxheG3yr+llhEvVfGTJKkg70xPkW/qTbpwqo6+KsQaeDJViMIjRKJSBFC1KehAlcVFEIrKDkDQ9Vo42xZKNNJOiQP4wqp5s6hL5kHrSlUMMImHpA4taKhTnJRxOqYw8UarEEE0WBlbshZ9eMk46MSQsivDH53rjY6UAVGRDzZQoIIix9SlwcEYuf0gzwu8fUWD7EPKMw2InhfFXRJWRJ0qVEGKJSb1dw2UdSqJITBThEAVPLNVbW8blJUA21kyIAoIY35AT6oUymZhD+oFWl4G9Y6WTuFIAKqqMPFGqRBBLJN3TCki7fUCim0SQkCjCL27qyXWlo02xZINNtyhMQZxcyvkOqllzyBQwLB/YNz5+j6Ey8kSpEoGSnHSv787nmNh+GHGLAv2aAgem66UccXsJkI02naLICcJ2YJtgEQY8BvagUInBpMrIE6VSAArCW0yqT6GLG/gk4x9siVsU2pPLigrYeMiGmy5RGDmEzEH4io0zzcEO4Gkc2DOG74exyk8lCFBl5IlSJYCKyMLwr+pEkQ+PjTOtGvGJAl5i13jdIJWGXxnxnijGGngyNAUxqi4n1dOzI4dAKbfnLYVfXSbt/j6OhbeQdmcrhR8eosjbG0S+Twk97eyA7C+ybRAbY221IECVkSdKlfFXRAzRTsyn4LlVfI3jWw4QlyjCTy6RZ1LTBMMmk2zAVouCBSHzEBh2lR3+XQzNT+E31yh0fi4FtnemwOpm5F/WkPxL6pN/aQPyLy+gwNoWFNzTg7Rrq9nwsPDJuWs9MCqF5gJivE4QBTjyB9m+OBynt6haFEE/e4lxukE7QhR8zKE1yLe2N3/Jp8ZJuhOR4pekXVpEgfWtyL+wLvnns9AXVMLF/MTb1pG9yFYiv3O9Y/jFNX3hklNEAW8xoR4/eFbzyVVd91alKLSnV8g7lXMJlHM4RBQo3Qg/u2qcoTsR+XCXgocG6R5BxFAnPvJrAysaU+jsDAm3HAkO81B+7p1dWD7xVhl5olQZflVEbrGSc4vPL4yTrBiViyIUoOD+6RzD12JRsIHbLQo+nmdyEwqhlsnFi4IQAgX3/8HeQTdypfFXRryHxRQ6PZlzjY/GUR0GtEo9u0q27yojDJWRJ0qV0cdBL+cWoSvsZauYt6hUFOFXd/R5iYHf6YKwUxQ41qh6FDyx2NULgyJsxKETY1kQ9ZIThEm8l3MP7epKdufOvB4R70cK7BnNBhmVeKuMPFEqDD4+GvMWnvfGGapRsSg4mQsWraSSYXW+egnbRMHHGVKT/Jv+cnf5BsKKuzs5/GmUmiBM8jECG9pQ5NVl4wOch/C7B+TnJNcRosC8hcxyHzLOTo0KRYGOFr5lv7KXMHIJO0XBx8G+eGHOb9yMiOeNjCJZIgghH4dDsNDpKew9vcanOAwcqmh3DpF3emN9/kJl5IlSZfBxkT9/tF4TRYGKJ3orEAV/kZuH+AANynoJO0TBx8D679C59eK93IzwkxPkX85eQkaTVEaeBOEtNrWlyMcHxqc4EAEvBQ9P5/uILQYsEIbS4OMkhmdRQfu84oEatSiCPvJvHV5WDHaJYgjHgQibZE21ixEOkXZuttqwUyILbGkD0h7s4w9x7przCOYvVnWxXxQYnh3P3rVoaYUPWaUoJMFG+3y7RcHvR2kJJg9dj6CHggf6WRg6RXFhXQpdnO9sTyrrcHaRd2pB6sJQGXsiHPWD3uTgyyvj5MqivCj45IOnV3OCXdt+UQzHQnS+2VnQhibi/UCBHZ3TIwpm6OQ4FkXQ+DRnAv1sAzuHsWHWKm/oiVBl6InQTLjvHzfOrCzKiQIn7l/dq3wukWlR8Ht9iztS+F38hVxOBibaAts6pkcUfMzQ8VGOFwUQfnpRn9RLxVuoDD1RYhHSIQxQlG+gVk4U2qML5J3c9OsMth2i4PehS0jo/Ea+itmxvhrzE8GdXdPoKSbwtXJB6x5MCB+ZoS9lVRl8PFQZeaJkb+Fb2l45w11WFAidji+hkqEcOtnpKfh9vhXdKPLppXFiWYCQj0KH/kpbTqFdXsL3zx2jc+HXt8m3sE3y3kJl5EnQO7mBDBfHoowosAuOf02fir0EWM7oqyKMPIoqEUST3yNDsBc344z0E8sGsMGGLi0SA1YadtKsTYFlBRR+nFqvo4wC3uLEAn3tBUaDVIZfGRUGnhQ5hAoemMjhUdmws4woMDkmxX/RZR2xLGf0VRGGHkWVEKLJ7/Et+4XdWvq2b7IL4RdnKbCqibXego8V2NaBIl/cVTEcfnVT9xaoi1IZfmVUGXgyRAi1nK/d57IRyVdRYMjs3AbyDI8p64hlOaOvijD0KKqEEM2RdfVcgs8n2yB5xaGBFoqCj7OoHmkYjnVo/VOF4AQ3eNjILRINo1QGngxlFKqR9LKKRqkopMHZlqFU8ue/2VOA7C1MIpzKROk4v967oD2FLWiS6yS8fP6aHj14SpqmcZhzTF9IZIUw+BjBbT9T5P0d45PchfCTi187DqqMvyKqDDwZInQbV4eCJ/mhElV1XUYUoZPLyb+uH/lWdiff4k7km9uOvDNakGdyU35zAR+oPru7uvocxuCaYsRCrNsW4meDpaIAYfAGVWIwOYxjvGML+KmXPQ2QH9x7TAN7jaYBPUbSm9fvZRIvdH5O4usoYomwaVVT0u7uYOtyaflLoIQCO4aycSY4b6Ey8GTJwvBv7isVvSbKhE8U8Mg8BRJu7OqDfpyRtw+ltU348QW+ASdJu76PQhc2Sf+nwIEZ/KVGkX9Dfxkt8vFT3juzpQzpesY15A/MZxFxOIbRLBERG34ZsURx0Hfsypq7fvGQiQhfz1vX71L3Tn9Rjf9uSHX+3YRWL91EoZDG1/e1zCtgJV1SwsB7lrPbv7xUROZmaLcPyihQQgm3yriTJWqhFraWHMdEmUQ7PrB44GrwdMITHbPNQR/Hy1+kBxBaVcpm5xDR7SMUuriV3dMyaY/j3zac/Gv7km9pZ/LObSslHJ5JjWS0CeIJ8N+T6ejmNCBMOn/mCv3cpif98F8N6bt/1Kfv/k99atawHR05cIr/HtaFcXrK1zLyeMQhr2EPsa4FadfW8HV3f0sfdAHBiriEQiiVcSdL5BVTG5B2a69xRkmJIklASNj83F8i4gm/fyLNj7UHRex99nJIsUE6mbt5RR0QCobo6MFT1KawK33/nw3KsXWzznR4/0kK8uso8IXCd7ZRcOcvepMCEUetryIpJf8OYljRmIIHB1D4KSeGmjVbWdkO6Tq5ILHJPJVxJ00W41gO24/PEfsEMieKaoBAIEi7th6kFo1+Eu/w/T/KiwK/b9bwR9qyfrcujAiHU5+eSDOC0JGhFNzSXhJxeBCIILCmkILbO1Ho1AQKPzxorMvOrpE57ckFzl2bxO8tlMadAlEguKU/X1u9EjsnCovg9Xhp/art1Di/tS6IGDGY/DeHUQ1qt6C1y7eIVykFPCR7jsgXzuPeXKPwy/PMC5zT3ZSuH27PHSoD2hT51/eMf85CZdipkEXhW9KWo5eHcj45UViAz5++0KI5q8XYYfQqMYD4W5P6bWjD6u3kYRHlYAC7Yp1azGEMQqg4vIXKsFMi5xXTC0h7dEZOJyeKFPHuzXuaNmEB5ddoVqmHwN9aNP6Z9mw/RH6f+0vhrYb24JS+ZDWeEEpp2ClwDHNCHQpd3iQeOyeKFPDi+Wsa9fcUqvNd0wo9BMTww381oLYtfqUTR87IkGwO5RH5+FS6+MUVQqkMOxVCFMzg4amszgDlIRa+ffMe3bx+lx7ef0LPn72iN6/f0ccPn+nLlxLylHjJ5/NTkJNIDDVi/L26A9cA12pAz1FU65+NKvQQIoj/25C6/NiXLp2/lrt2lUC62qMdjkzkVeEtVIadCiEKtL/Z3JfIX0J5Vy7eoLaFv1Cjeq1kVKRV007UvtXv1LV9P+rTbSgN7j+exo+YSbOnLKFlC9fTprW7aO/OI/LUu3juqkxQwUBesJjevf1AXz4Xi4g0TFJloRFEwhG6fuU2/d5poEzKVSaImv9TQH1+HUJ3bt7nNxoHyKFCYKNJr7kBjEoMJlWGnQoNUfiWtqPIl9eUt3fnYYmHv/2PfAkBTH77H/wvyD+bIQCMoNa/GlHd75tSg1qFLKTW1LygPbVp3pU6te1NPboMokF9x9C44TNo9tQltHLxRtq2ca+My587fVkE9OTRcxFPSYlHhiQzLpxwmL/4W9IeX6PQ1cMUOredQme28r87KHT9CGlPb1Ck+IO8LhZh/t2ZU5eoY5te4gEqE0TtbxvTEH6gPH6U23cvXoQfnY1vaFZl2KnQEIV3dlNZ65EHw63Nhl7RDS7Df+g3HDSF8+3/zv9KQ1jf8etqfMMC+n+mgFpQk/w2nGh2oHYcW8ML9es+nOPxqXRo73GOs9Nf6xQJeEl7eIkCu2eRd05XKhnZjIoH1qfifnWo+A9m/7pU/FcDKhldSN75v1Fg/3zSnlwjNJg2EfAHaObkxfK98D1V1wgz1/V+aEoTR82m169y++4lApQVlTZmVonBpMqwU6GRU3inoCvKScqbOn5+pU+9ZAhRRIsHHgeC0RnlffhzYWTIV9KGsCZPf/+G0VQyrDF96VmDvvz2LX3p9i8m/4ufTeL/+D1+7lWDSkY0Jf+2yRR+eVeOAzx5/JyG/TmRPUF5YeD/8KDzZiyXnMy1gPeWyoLMevFIoITj+v5s+PaIwjOhLoWubKE8uHgrBREv8ZnwUFs37E5bCBXxlVDw9BbyjGtJX3p8rxv77+C/q6YplB4/kGdKWwpe3CPJIPDq5VsaNnAS1fxnQem1gyAa5bem1cs2U0mxyybaWPDoXhh5fVVmzbVbm0m7uYG0u7so/Oy0vidGMAPzKizE4JGZbKhVVM2qDDsVmqIYV5uCpxZQXo/Og+QJHmu06aY8VWu3oNMnLxhXxFpEPJ8osG8+h0T1DTEoDD9e8vuLhxRQ4OgKivh1g3/KHqNH578k1/qehVHIudX2zfvI73dRTRJKTD4+IO3qKgru60OBdS0pgA1jUL0LLmkg5elYsxEqmsQCOZX2IsTQ5c363nkqMZhUGXYqNEWB5an7x1MeKjmRD6gMN52EEFEjdPeW9e0eUXSInKC4fz09JFIZeqKEMFhgweOr2WPohn/h7FUq5O8gRX4HTmYkN7IMHKpgLQbqqvS1HShENIoRsaIPNAsS5/HvF9ahwNpCFscU3XOkKbTSHhaRdzKWHbCRVkSVYafCUlHU5HD5T8pr3ayTPaLgz4QgMZRrKTgUCJ7dKkmzZYIw+du/qGRoI447D/AHRaQAEKXgF85ekXJw1wDbAWCh04rGXw1fVapehsbrFtWj4O7f9E7naahoDr++Q95ZzSpPtlWGnQpNUWCuYt1vlIe5CTvCJ3xm325D6cP7T8blsAba81ucMLVOPWSqiHxcz4yOfPP04jHkQ66aj/F/phB62pql6uWMvyrq7wnu6Cw5iNVAEwHfora2icK34mfKw1yDXTnFyEFTZMbcKiAR9m+ZIMmx0qCtIMTWu6aEZ65r54kmz7c2sYdIdX8MQxgHB+gVvBZC71D5i32iWNyG8qSy8z8yO/okw7GcoKKQDiGIVcC8QsnIpunzEiY5LPNMbF3qLdyCyLvbFNjangWBRUsqY0+AEBUn4to1bMVr4ZB60Ev+zf1sE4V3fgvKw2x27Hh7uglR1PjvAim3DlsVi3MIEziwkIp7GfMQKmO2jJx096lFwdOb+XNdkktoAQpdXCA5gfmkT5ksDOk59dnCfr8h7MY7wj5RzG1OeZhxtkMUmBVeswxdAK1BpOQjeRf1zIAgTH5LvtWDSucu3r55T2eLLtGp4+cdw6IT52WyUebiOMyRkaaUwqZYwlvUJ+3ONuseDlqQw7LJajGYVBl2KowWxeymlFfn28ZSmqAy3nQRIqzzbRPavG6XcSVSR/jFHSrBJB1mpJVGbDGRcE9uS+FPeidDGCDWSzSs05IKHECcB0pr1q3cJgMB4een9VzCKi9RSqPjuVVrxrUQBY/NZuOvZAJPZdipMFoUMxtTXtx1TxYSooCHwmSXVQjdKaKSwQUZFQXKRrTnt+XzL1+4TgV1Wxk1YEZdmI381//Kp7rfNaEdxjXWbqw3QieVYadA9jzBHV0o4rNopylpZDCfDdQmUcxoRHmoerVDFPVYFOYNswKYOyj+M9/6uYmKyKLAXIh2/7x8/r3bD6l5w/YZD0UrIgZP4C1Q4g9omJdQGXWqRF6xsY2sLbcEmGcqWqLnFOb+27G1UCrDToXRopieTaK4tJeKB1g4g10lWRQDOZ6+e1o+H2tKMLvtFFFgmB3D7XoZTYRCZ6apjTpVQhTrWxuz3BaAcxPt1n6Zq/BOb0LeCflsrOw1SkWClXn8s8q4k2XWiuLaITHSjIpiEHuKhxfl8+/cui/NzhzlKWq3oKMHi+T8ZOTJ8m0AmBDF5nacyJff/CRZoIVr+M1dvranOQLYSsHjc6W9JjaS9M1rQd6pDaUHrBi0eBGIJAWhxIrC1pxik3Wi0O6dkxKMjOYUI5pS+NV9+XyUeiBcQU6B72c3cR5o1Yn+UkD49la9yE9l2KkQOcXu32WmPG1AEz10oPz8ShfLg1McGWyi4JHpFNjSn3zLfiLv7GZ6+02IZYy5pDWGFQknWhTIKXDh7Bh9qv1tE9q4dqfxrVMHJtI8k9pkVBTeGR30VXqMMxymdGjdg1o2+ZlaNu1oP5t0pNbNu9CmdbsoHI5Q5PVlqXiFESuNOwXK5vaZ3loMQ8Do6If+x8VvpWZKu3eMQufXUPDgJPJv+kPE4ptbqO/IOqHeV7EgBBMaYhkDGqKY1cTeeQo0HLYKqIz1rfhTbcBp4Xd84cdyXKKXehR/KaHHD5/RI84tHMMHT6W2TCqzvO8ouLenxaLgYy0voPCjw/wBTqn/4vPAgrBACXuWFxR+cZ20O4d1sRyeSoFtAzkM60y+BS0lVPJCLGNZLPAiI7/XRZFfo7ktosCM9sLZq6RDiFUInlxPxX/Ulqe42pAtIkae+tfTq2XdUgzIhoKFQ8kXAiqI0GlPD2kW7QqwN5Mw7BOL5eVNDsNOyEq74Il5FNiN5t+/kn9jL8pDMmZX7dPUcfMpELCuqC785pHUJGVCFN7ZnfjiWlz2nmZEvjxnI7bIW/AxAisbc8iyhw/sklKXiiCNv4slDEOVbh56n9pVJTt84CRrl27y0zBwaAl96V2LjTdNwhAvUZeCRRz6uXA74/CTE7LCLiVh4L2ctMuCI/8X48jZgzy0qLFlkRELsfcvg+n9u687yFiB8Ifn5J37i9qgUyYLrft35Fs+QNrkIHBCH1n0unINtIB0OA+saZ6cMPCeRfkUPDpUPE82Iq9Ns872iII/86fWPaQjoaXgGF+7XUQlY1pYH0bx8TDCpT26Ih/lKfHQnKlLafKYufTqxRv5nSsQ8nHYs5sCm9rJMtP46qH4NQiZOLEOnRqfNkFgSa/Xy+dnYzfKPDT2sstTYLH/bXTPsxooKruwi0qGNzEm81IVB78fghhTSKEbxyRswjDngT3HZIIMcz1oAofRHteAQ02snAsdG6EP1WJiD15ARYhiSQOpsoWXIb+1qyWjgfaig/qMoYmj5tD6ldvo5NGzdO/OQ2mghwVpwSDH/2kWS16vroOlgExluOkkcgo0STt1/JxxKhYjFJDSD6mchWEn6zXwvu7fS0Vs6MZxMSbg2uVb0jQZgwb6aFpD+q3jn3Ttyi3pJOgWRALFFH5+RspAgru7UWBDawmtZOOYdS0osLUDBY8MJu3WFr2+KZLeBtFrV2yhWv8soB/YRtCnF+t9mjZoK3NAA3uPls6TWzfskXkh1Juh3RCGw61sXJ03bMBEuakqw00n8ZkoMdm0dmf6jAjDkPfPk29xb+kEKAYerziM1xYPyCffyr9Ie3y1dJQFDakhADRzK/1O/2ggI2rtW3eXp5urOnsAGK70cJ70jkX94pwIJfz6CkU+P6FIhjaMQfOHSaPn8HUtaycgHqIgrjFakjas00ImStFtcuiACTR3+jIRCx6yKLl58+qtbKQDj54o8mZMXCgtLu0QxvcyLDvP0mHZ8ojImgfMYXhnddKLBjlZLhWIiCTqZ5A9Q/HAfPLO60bBc9v5CflePxK7bXiIXzv0V14z/QY2kNlk7ENh5VLb6gB0ue/ZZVCVtmiKBCE4Qn/8DE+NiWiU72NGH53eEdKiAyUevCePfQ3DpAF4JTlLHla/QXmZLvUA8WXQjC0jozecZ0Q+vpLCQf/WSeSd3YWT8UKplyoeXCBrIzxjW8jIVWDHNNJuHuen5NtS7wDgQqKWCK0xKww5jZahjfPbcCiwVRpJ5xAfkJOh5RLsQnltK2EZoTD1Y+hiQSkTehFgpBWDO/17jpT+ANhRCl4dXeHR97e4uESiljwki/X5JidzIqkSn4mGaI8fZjJB5acDJms8nyj87ilpz27KaFL4GYcN759RxPtFBCSvUwDx64rFG6Qsu7Jrhr/hus6ZtpQ+WDzsnK04cfSMFFVaaYu69y4bguFnhGHIXZCzYMs15Iejh0yjd28+UB5WjCGRsfJEEiFO6tghfU2CW+Dz+mTVINZP4AKrvheIv6ED+dhh060fes5CLF+4XhpXV3ZNrSY+C0QYhn1a3rx6R3nPnrxglfximyjg3ubPXJFUQmQnsLfG8cOn5QmDp05FNxLXFUO2/XuMpPt3H9k29u50oLJhcL9xGRVENPG5f/w2TOae8hBH9f51iK0n07PL3xKWuA0wcOwEVZp4V/BgwXfE3zFSgte7acg2U8DKRexdYsecGYgH23TOM7AtRB6GwaaMnUc/8E1TvTjdxJO0ecMf6ea1u8blcR/gAQb2GiVDzJUJA2Pv8CzHDhWV3UM7B9ntCvNWdkQsuDfw5hhEQcQiu6NuXrdbRqBUb0g3cULoOrFxjXULjuzAy+evadzwmZJD4DupviuIvyEX2bZxD+cmLmrbn0Zgh6hZkxcbmwepr1s6CSGiLdC5ostyPiKKi+euSdZf2c1MG42LMLjfePdtdhIDLOjBLkaVbTKPawxvgnF05HM6MCLmlxnj8NOTpN3cRNqV5aRdXUnane0UfnGeIiVvZIItG4HhUFwPDKWqrlm6ic9Fkv3E2J9QRPH65Vv5pS2iYOKkMOF164Z7QygTqM9BAzIM88UKA9cXsSu2IEDztCBKE4IemUEOnZpIgc0/UgANyxbXl0pUobFxSmB7JwpdnC/9YC3t3eoAYBYaDxI75spAzDkN6DmSPn3U15mLKODGB/Uda5soYDx1OITasHpHVozOIBzYt+uIbL+sT+bp3xOC6NSuj8yK43uiB2vo7HS9jBtFd1J8VwHxt0X1KLCpLWnX1lDEq8+yux0o8MNiszIlMxkkbN4cATVLc0QUSLaXL9pANf6nQPnGTBDC6N99BH38kL4KzEwC1/T0iQviFcyNNjFbi92PRBBvrlNwX2/dG5iVqPGQX4stuEInx3G4ZV1bGbvw7OlLW6MU2B0mWQ/uPW6ckSEK4MzJi7a6MJwcVgGeP6OvVcgGwPhvXL0jpSwoYEPBmobOGu/vSFsY3dATEIRJiAgr346PliI+NwNtjlCzZJcoELrDo6OI0ESpKDDj2r7l77YlO2CNbxrStPHW7lnhBDx68IR2bztInz8VE3HYEzo6VF/ck4iHiCXeu7QBaZcWyaIhN+Lzxy8yYaayhUwRYsSe7lhBaaJUFEgQRw+eZptiQXw2VPvwnoX7HTgEkiuhlP3WFlm9lpIgTCKUWt9SRqfciBNHz1ZZQ5ZOwt7Qamnx3NVlctlSUQCYvEDCa9dJgpgAQw2MqzZWjBOR4lcU3NXNGkFEMXRqgoxiuQl4CI8ZMr0031LZQrqJVAGl5rHbVpcRxfWrt6Vq1c4QCif6U+vuUWP42YKINA3zL2MvkUweURHhLTa0psiHe8bnuAOXzl+XUm5bbY3FiIGQN6/fGWelo4woME6L+MpOT4EThbfCOo+Iy4oEK0U4RKGzM9SGnRJZYNhN6D76L7njeqExwaTRc+1b3GYQXmra+PmkxSxlLSMKYOWSjfqiIxtPFqL8uU2vDK+zSDOCJRTc/4floZOQk/bQhXmSs7gBWK5gd4d22Hf9ms3p2OHyyxbKieLqpRt8wj/ae8L82RDmsgXrpEQ7GxDxfqDA9s7pEcV8FsWJsa6Y6fZ4vDRq8FTbJutMwr4RpqPEJBblRIES7n6/D2clqQ+WKeKk2zTrkp4WODYA8wmBbR3TJApj3zkXiAIjTgX1WtkaiYCoCp85eZFy+L+cKDA0hYXedlXNRhPT7+jugBjU7Yj4PlJw5y/pEQUTtVNOLxjEDrJ9ug1R3utMEg9cjDoVnSg76mSinCgALPhA+xC71YzPxzj2iSNnjTNzMUI+Ch3+mw04HTlFPdKuLGPlOTenwBD7+lXbZfmxnaE5iIbiPTr/JSJVQSkKv89P44fP5LhPfdBMs3unQfTSTW0pVWCDheHCgJWGnTT1VpbhJ19rd5yIm9fvygIruwVh5qsYUKpoBaRSFABaf0ibfru/BHsLtCjBrKPbyz/CLy9K5z1LQyg+FsrKnVwciDx1NCfXCIftjj5QJo5lCpWt9KxQFGgahcZUdosCxDlgRKzouDvLGUxE/J8peGQIG7NVouDjLM4n7fJSjk+c+cBA2IRVhvVrFjoiHMeo15ih06UhWkWoUBRYq4r1DXZs/1URf+800PUz3eFnp6RHqyXego8R3NGVIh8fGEd3HtBbt3WzzrYLAsQ5YNkp+ktVhgpFASDhbt8KlbPO+EKoi5oydq67l62iDf7lJeRf1jA1YSBsWlNI4QfYYsyZdWJY0dm/xwjb5ySi2afb0DIVsSpUKgp0cp4/Y4VUEjpF6WhZicJFK7tMZxpYNYcdRZPbf45fD0GsakrajQ2OLRvHJB327qjzbeZ331URtoOmEujxG10Rq0KlogBu3bgniYlTQihcYHTDOHXsXJVfzsnADLd2Yd7XxLtKceA1teTnwMY2+j4RDhUEaomwoAqhilPsBufxy0/96O3rqpfxVikKjPigSRR2M1V9mF3s1LY33bru8kYHaFrw8AAF9/XRGxaYhm+KpJT8u4UsBhZQ6NhwCr+84NiJOjyo8MBq0aiDI6ILEIKAl9gsfZ2qDjWrFAWAMebCRu0d8yVxHjW+aUh9fxsma3xdDc4HIiWvRByhookU3NmVAutb6ZumYPMU9grYljd0bjYn6UUUSeMuQlbgOifWyENV980u6l6iP72Is59vXKJAd4rp2Mfiv52TMEEY6Bo9bOCkCmcm3QUOBUNeinjeUOTjfQq/uU7htzco8ukRh1rvOCZxfuM0dEqUzWwq6a2baZpeAqVL8YbbcYkCgLfAUlEnjESZxIVHl+oJI2fJvEoO9gFl/n/8Psz2NRKxxLmg1y86OMaLuEUBb7Fg5gq9X6rDvjTabqJwMCcMe4C5oz97jaKa7LmdZhv5NZvTjs37ExqUiVsUwIO7jyVedMqIgkmcD1brQRhoXZlD5vD40TPZoBGhrJMEYRJ554cEe4klJApk7utXbauyibAdFGH8uwmNHzGr3JrbHNKD+/yQRIsajEw60R6wbgN73SWKhEQBIKnFfhKqE7GbuDFoqY7dMp881Jvl5mA9EIqg9SeSaicU+cUS54PcBg9IryfxuZyERQFgBx+sc3DaxQBxTnDl2IjmxtXbrp7gcyLQbxXzENhQEeUbTrQBeIk2zbvSjWt3jLNODEmJAuqbPGYu1fwfZ8aROCfcMKzBlT2tcxukWALc922b9lKLxsbEnM1LllWEIBDer166OelSoKREAWBMun2r7nISqpOzm7hpIErOkQdlZFviLAbytDnTllm+e2k62LfbsJTyyqRFgaQbzXEr26DECYQw0MoE8SV6uubCqcSAOiaUf2P/Binug4dQXGcnUH8ItqNzRZeMs08OSYsCQAku+s86MdmKJs4NeUbX9n/Q0UNFlS4wyeErPrN3RWFfm8JfZGmy0++xvkJzTcorNFMSBXDn1gP6sWU3R18wEOcHYochlDQ/f/oy5zUqAFbL3eX7iv5MZiTgbEE0kE02+3YbaslwfMqiQBi1d+cR2X/B6cIAcYMxK4891rBRhxu3Kk4X8JBAVQDqhNo07yIRgNPzBxDniOUN2BDHCqQsCgCjEjMmLrK93Wa8NM8RQsaCerRx9PsDxrepnigp8chIHXoJO3FytiJCENiJCO1zrBpltEQUwPOnr6jPL0McVSFZFXGeGLrFEOPsqUskZIhttpvt8PsCdOn8NRo7dLqEluZ1ib1WTiXqrcYOm0GfqlhimggsEwVw8dxV2dfNDS7XJOJRVP5iyS1ChgWzV9KDe49lg8JsBnp7YVQJ9WJoiV/jG3eESrH89af+pVv9WgVLRYH8AttYFdR1/lh2LPF0xDlDHCiRn83JOPbrQFiRLUCHlk8fv9C505dp3PAZMnxp5g1u8g4gliWji+X5M/qG8FbCUlEAPq9PnrZ2bu6XEtlz4LxhLM0KfpQRmMP7T9CrF29E9G4EvB6qWXduPSAVrSjRcfoQa2WEiNELFiXh6Qh3LRcF8OH9Rxr591THdAFJlqa3Q6dELGdcOGuVjHBgcxs8dZ2MsBaW4k10+caqSZS8oI8r7ofbvHg0ce4YCFg4exV5kij2iwdpEQXw5PFz6v3LYFc/kUziRoAQeZP6balHp7+kjSe2N3735gMFAgHb5zzgxTApiRVmGEWaNWUxdW7XV0JZeD1so+VmMYCwo1r/bCSDAnjwpgtpEwWAJawd2/ZSfkE30nzK4l+MemBIF9Wi44bPlEK5Kxdv0gs2StRZpbMIEQLErC1GXLDq7QKLE0OSWK+O2We0qDSrDMzzVX0fNxHfAd+pf4+R9PJF/EtLk0FaRYGbd56TOozqZMONiaVpdPgZczTYHL8Di2Rwv3E0b8Zy2rZxL506fp5uXrtDT9lzYmIMk4VY2qtpmoRgsR4G/8dTHxWeGCGCwN6+fkePHj6VNQwo28fk2sxJi6UeqV2LX6W/koSq/2mKQH2+biWu8w//1VDWb6AQNd1IqygA3OBjh05/LTdWfOlsoCkQM0yp8U1DWSKL8AU9kLCM97eOA2Ut88i/p9DkMXM4xFnC4llBS+atla3MULcDMc2ctIgmjppNw/nJ37/7CBl2bNeimwydIr+BADEfhKHkbAiLKqNZwtGxbW95KGQCaRcFgKfivl1Hbd9LL5OMFsm3/9sgG7HZQhKThljGiTUpeMrrLKAa/H8IytwbRHUM/C6bHzCx/JEfCBjgyFTelhFRANjQcff2QyIMJ/QWtZN4+sGoq6b6/dWCxndvyzkSBjQyOdqXMVEASA51YbSTm17uQuSYI9N8GCBfKjpxPuP7qWdUFECpxyhgj5ETRo4xFEFwvtS+5W90tuiS5KSZRsZFAWDx+4E9x6iFsdlkThw5grADDCD83Kan1NHZVUFgiygAJN8njpyRmNG8ILEXKcfqQ9x/zEP8+vMAunH1TsaSahVsEwWAUoQLZ69Qp3Z9HNeDNMfMEQMvWPiFeZd7dx4a1mEfbBUFgCfC7Rv3pTmv22ulckycuN8oHsWaiOfPnLGtgu2iMIE106hpQbFXdZnLqM40h6XRMmfBrJX0McF+r+mEY0QBoPp06YJ1Ui6BC5bzGtlJ875iXfWOzfvI53VWdxVHiQLAWul9O4/I7vxObcuYY/LE/cQs/q8dBtCZUxcduaGn40QBoMUK6lz6dR/hmmYIOVZN3Ees6UD+8OjBU1tHmCqDI0Vh4vWrt1Igh5Vi3/MFzYnDncToEuYfUBS6YfV2abLmZDhaFADCqUP7TshuqFjDkEvC3UXcL3Tuw6buKOpzw/7njhcFgNoXdNhAOTXWDuQ8hvMp94i9AwpAF81ZRa9evjHupvPhClGYwAKd/buPiteQOY2c13AcIQbcF6wlgXc4ffKCLKpyE1wlCgDJGVaxzZ6yhJo2aCtPo1xI5QziPtT4pqG0CFq9bLOsGHRqMl0ZXCcKE8g1UEWJlWxom4gbkgur7KF+7RtQo/zWNGbodFmb7+ZOi64VhQnshrp98z7q2r6fJHQ5r5E5mrPS+TWaU7/fh9ORA6eopNj9zeNcLwoAJcboagGXjUk/s4Yq5znSQ/PaIm/o0v4P2rFlvzRlcGOopEJWiMIE1mk8fviMFs9bIyXp2CkVniPnPayheS3rftdU1jxsWLNDRpXsWveQLmSVKExgLPzB/SfSsAwruCSsMhb8q252jpUT1w0jSpiNRpiKHlMvnr3KOjGYyEpRmECy95DFsWb5Fvq1Q3+JfXNhVfzEdcJMNCpZUdqPPQ7R7M3pLUNTRVaLwgRW+aGr3L5dR+iv3qOpcX6b0g56OZF8ZfT1wB6BhQXtaczQaXTi6Bl6/+5jxhsI2IVqIQoTSASLi0voysUbsu9d+9a/y86pemMxo6mY0UmiuhACwPdGP6ka3zQUr4ANM5cvWk93bt4nr9eXNQl0vKhWoogG8g40I0aZOtrttynsSvV+aCaGku1d90BdCPlSnt+gVqF0JZ8ydp605cRIUrbmC/Gg2ooiGljkgrXBO7cckMkndKTDhCAMRpLMLAmvzNAIoSM8Asplpo6bTwf3HJMu8cEUt9rNFuREEQOPxyvbRWG/bZSSYDEM2u9jY3VzWwEIxRSL0wRjnlP0Of7AYRHWQaMXbZ9fh0jPWtQkYQQJTZxzKIucKCoAwmg0bsN2WLdv3JMJKuwPB5EUNvpJQg7MgyAfwXCv3udVD7syIRb5DBg+PtvoNYvPhnfDwixsE9CyaSfq2eVvadh8YPdRqTRGUaUbyrftRE4UcQIiwSgW1pFjF9XD+0/S8oXrafTgadS901/UlnMS7C6KxB3tWiT0YsM1DdYUjdkgOV7q74k+hi64Gvz0h/FDnM0atKMfW/5GvdkLTBg5i9Ys2yI9tTAcDRFU5/wgGeREkSJQFo3hSjyF0dUOnQ/XrdxGc6YtpdFDplH/HiNkazC04kf1KNYXYEgYe7Y1qN1CcpdYot0+/t6EX4dh0dbNOsvmMNifYWCv0bJJzPyZK2jjmh10hMWJfcAxk/+RBQvvlkNqyIkiDcAQJozTy/kJPMvbN+8kfn/ET24Mc167fFOawBUdPy9P9Fgi3sffr1+5RXfYK2E9M0bK3r19T58/fZFhUpS05JAOEP1/oR4vzuVXBOwAAAAASUVORK5CYII='

ss_path_b = os.path.join(config_folder_path,"my_screen_shot_before.png") #before search
ss_path_a = os.path.join(config_folder_path,"my_screen_shot_after.png") #after search

fullPathToStatusLogFile = ""
global ai_screenshot #AiLocateImageOnScreen
ai_processes = []

global HLaunched
HLaunched=False

def timeit(method):
    """
    Decorator for computing time taken

    parameters:
        Method() name, by using @timeit just above the def: - defination of the function.

    returns:
        prints time take by the function 
    """
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        print('%r  %2.2f ms' % (method.__name__, (te - ts) * 1000))
        return result
    return timed

def read_semi_automatic_log(key):
    """
    Function to read a value from semi_automatic_log for a given key
    """
    try:
        bot_config_path = os.path.join(config_folder_path,bot_name + ".xlsx")

        df = pd.read_excel(bot_config_path)
        value = df[df['KEY'] == key]['VALUE'].to_list()
        value = str(value[0])
        return value

    except:
        return None

def excel_is_value_exists(excel_path,sheet_name='Sheet1',header=0,usecols="",value=""):
    """
    Check if a given value exists in given excel. Returns True / False
    """
    try:
        if usecols:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols=usecols)
        else:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header)
        
        if value in df.values:
            df = ''
            return True
        else:
            df = ''
            return False

    except Exception as ex:
        print("Error in excel_is_value_exists="+str(ex))

def update_semi_automatic_log(key, value):
    """
    Update semi automatic excel log 
    """
    try:
        bot_config_path = os.path.join(config_folder_path,bot_name + ".xlsx")

        if excel_is_value_exists(bot_config_path,usecols=['KEY'],value=key):
            df = pd.read_excel(bot_config_path)
            row_index = df.index[df['KEY'] == key].tolist()[0]
            
            df.loc[row_index,'VALUE'] = value
            df.to_excel(bot_config_path,index=False)
        else:
            reader = pd.read_excel(bot_config_path)
            
            df = pd.DataFrame({'SNO': [len(reader)+1], 'KEY': [key], 'VALUE':[value]})
            writer = pd.ExcelWriter(bot_config_path, engine='openpyxl')
            writer.book = load_workbook(bot_config_path)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        
            df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
            writer.close()

    except Exception as ex:
        print("Error in update_semi_automatic_log="+str(ex))

def gui_get_any_file_from_user(msgForUser="the file : ",Extension_Without_Dot="*"):    
    """
    Generic function to accept file path from user using GUI. Returns the filepath value in string format.Default allows all files i.e *

    Default Text: "Please choose "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Please choose '),sg.Text(text=oldKey + " (ending with .{})".format(str(Extension_Without_Dot).lower()),font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue ,key='-FILE-', enable_events=True), sg.FileBrowse(file_types=((".{} File".format(Extension_Without_Dot), "*.{}".format(Extension_Without_Dot)),))],
                [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                    event, values = window.read()
                    if event == sg.WIN_CLOSED or event == 'Close':
                        break
                    if event == 'Done':
                        if values['-FILE-']:
                            break
                        else:
                            message_pop_up("Please enter the required values")
                            # print("Please enter the values")
            window.close()
            values['-KEY-'] = msgForUser

            update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-FILE-']).strip())
        
            return str(values['-FILE-']).strip()

        else:
            return str(existing_value)

    except Exception as ex:
        print("Error in gui_get_any_file_from_user="+str(ex))






    
def excel_get_all_sheet_names(excelFilePath=""):
    """
    Gives you all names of the sheets in the given excel sheet.

    Parameters:
        excelFilePath  (str) : Full path to the excel file with slashes.
    
    returns : 
        all the names of the excelsheets as a LIST.
    """
    try:
        if not excelFilePath:
            excelFilePath = gui_get_any_file_from_user("xlsx")

        xls = xlrd.open_workbook(excelFilePath, on_demand=True)
        return xls.sheet_names()
    except Exception as ex:
        print("Error in excel_get_all_sheet_names="+str(ex))
    
def excel_get_all_header_columns(excelFilePath,sheet_name="Sheet1",header=0):
    """
    Gives you all column header names of the given excel sheet.
    """
    col_lst = []
    try:
        col_lst = pd.read_excel(excelFilePath,sheet_name=sheet_name,header=header,nrows=1,dtype=str).columns.tolist()
        return col_lst
    except Exception as ex:
        print("Error in excel_get_all_header_columns="+str(ex))

def message_counter_down_timer(start_value=5):
    """
    Function to show count-down timer. Default is 5 seconds.
    Ex: message_counter_down_timer()
    """
    CONTINUE = True
    layout = [[sg.Text('Starting in',justification='c')],[sg.Text('',size=(10, 0),font=('Helvetica', 20),justification='c', key='text')],
            [sg.Exit(button_color=('white', 'firebrick4'), key='Cancel')]]

    window = sg.Window('Cloint Fusion - Countdown Timer', layout, no_titlebar=True, auto_size_buttons=False,keep_on_top=True, grab_anywhere=False, element_justification='c',element_padding=(0, 0),finalize=True,icon=cf_icon_file_path)

    current_value = start_value + 1

    while True:
        event, values = window.read(timeout=2)
        current_value = current_value - 1
        time.sleep(1)
            
        if current_value == 0:
            CONTINUE = True
            break
            
        if event in (sg.WIN_CLOSED, 'Cancel'):    
            CONTINUE = False  
            print("Action cancelled by user")
            break

        window['text'].update(value=current_value)

    window.close()
    return CONTINUE

def gui_get_consent_from_user(msgForUser="Continue ?"):    
    """
    Generic function to get consent from user using GUI. Returns the yes or no

    Default Text: "Do you want to "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Do you want to '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow')],
                [sg.Submit('Yes',button_color=('white','green'),font=('Courier 14'),bind_return_key=True),sg.Submit('No',button_color=('white','firebrick'),font=('Courier 14'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                event, values = window.read()
                if event == 'No':
                    oldValue = 'No'
                    break
                if event == 'Yes':
                    oldValue = 'Yes'
                    break
                        
            window.close()
            values['-KEY-'] = msgForUser

            update_semi_automatic_log(str(values['-KEY-']).strip(),str(oldValue))
        
            return oldValue

        else:
            return str(existing_value)

    except Exception as ex:
        print("Error in gui_get_consent_from_user="+str(ex))

def gui_get_dropdownlist_values_from_user(msgForUser="",dropdown_list=[],multi_select=True): 
    """
    Generic function to accept one of the drop-down value from user using GUI. Returns all chosen values in list format.

    Default Text: "Please choose the item(s) from "
    """

    values = []
    dropdown_list = dropdown_list
    try:
        oldValue = []
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)
        
        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            if multi_select:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                        [sg.Text('Please choose the item(s) from '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Listbox(dropdown_list,size=(30, 5),key='-EXCELCOL-',default_values=oldValue,select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE,enable_events=True,change_submits=True)],#oldExcelCols
                        [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            else:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                        [sg.Text('Please choose an item from '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Listbox(dropdown_list,size=(30, 5),key='-EXCELCOL-',default_values=oldValue,select_mode=sg.LISTBOX_SELECT_MODE_SINGLE,enable_events=True,change_submits=True)],#oldExcelCols
                        [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:                
                event, values = window.read()
                
                if event is None or event == 'Cancel' or event == "Escape:27":
                    values = []
                    break

                if event == 'Done':
                    if values and values['-EXCELCOL-']:
                        break
                    else:
                        message_pop_up("Please enter all the values")

            window.close()
            values['-KEY-'] = msgForUser
            
            update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-EXCELCOL-']).strip())

            return values['-EXCELCOL-']
        
        else:
            return oldValue
            
    except Exception as ex:
        print("Error in gui_get_dropdownlist_values_from_user="+str(ex))


def gui_get_excel_sheet_header_from_user(msgForUser=""): 
    """
    Generic function to accept excel path, sheet name and header from user using GUI. Returns all these values in disctionary format.

    Default Text: "Please choose the excel "
    """
    values = []
    sheet_namesLst = []
    try:
        oldValue = "" + "," + "Sheet1" + "," + "0"
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)
        
        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            oldFilePath, oldSheet , oldHeader = str(oldValue).split(",")
    
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                    [sg.Text('Please choose the excel '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldFilePath,key="-FILEPATH-",enable_events=True,change_submits=True), sg.FileBrowse(file_types=(("Excel File", "*.xls"),("Excel File", "*.xlsx")))], 
                    [sg.Text('Sheet Name'), sg.Combo(sheet_namesLst,default_value=oldSheet,size=(20, 0),key="-SHEET-",enable_events=True)], 
                    [sg.Text('Choose the header row'),sg.Spin(values=('0', '1', '2', '3', '4', '5'),initial_value=int(oldHeader),key="-HEADER-",enable_events=True,change_submits=True)],
                    # [sg.Checkbox('Use this excel file for all the excel related operations of this BOT',enable_events=True, key='-USE_THIS_EXCEL-',default=old_Use_This_excel, text_color='yellow')],
                    [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]
        
            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                if oldFilePath: 
                    sheet_namesLst = excel_get_all_sheet_names(oldFilePath)
                    window['-SHEET-'].update(values=sheet_namesLst)   
                
                event, values = window.read()
                
                if event is None or event == 'Cancel' or event == "Escape:27":
                    values = []
                    break

                if event == 'Done':
                    if values and values['-FILEPATH-'] and values['-SHEET-']:
                        break
                    else:
                        message_pop_up("Please enter all the values")

                if event == '-FILEPATH-':
                    sheet_namesLst = excel_get_all_sheet_names(values['-FILEPATH-'])
                    window['-SHEET-'].update(values=sheet_namesLst)   
                    window.refresh()
                    oldFilePath = ""

                    if len(sheet_namesLst) >= 1:
                        window['-SHEET-'].update(value=sheet_namesLst[0]) 

                if event == '-SHEET-':
                    window['-SHEET-'].update(value=values['-SHEET-'])

            window.close()
            values['-KEY-'] = msgForUser
            
            concatenated_value = values['-FILEPATH-'] + "," +  values ['-SHEET-'] + "," + values['-HEADER-']
            update_semi_automatic_log(str(values['-KEY-']).strip(),str(concatenated_value))

            return values['-FILEPATH-'] , values ['-SHEET-'] , int(values['-HEADER-'])
        
        else:
            oldFilePath, oldSheet , oldHeader = str(existing_value).split(",")
            return oldFilePath, oldSheet , int(oldHeader)
            
    except Exception as ex:
        print("Error in gui_get_excel_sheet_header_from_user="+str(ex))
    
def gui_get_folder_path_from_user(msgForUser="the folder : "):    
    """
    Generic function to accept folder path from user using GUI. Returns the folderpath value in string format.

    Default text: "Please choose "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Please choose '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue ,key='-FOLDER-', enable_events=True), sg.FolderBrowse()],
                [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)

            while True:
                event, values = window.read()

                if event == sg.WIN_CLOSED or event == 'Close':
                    break
                if event == 'Done':
                    if values and values['-FOLDER-']:
                        break
                    else:
                        message_pop_up("Please enter the required values")
            
            window.close()
            values['-KEY-'] = msgForUser

            update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-FOLDER-']).strip())
        
            return str(values['-FOLDER-']).strip()

        else:
            return str(existing_value)

    except Exception as ex:
        print("Error in gui_get_folder_path_from_user="+str(ex))


def gui_get_any_input_from_user(msgForUser="the value : ",password=False,mandatory_field=True):    
    """
    Generic function to accept any input (text / numeric) from user using GUI. Returns the value in string format.
    Please use unique message (key) for each value.

    Default Text: "Please enter "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value == "nan":
            existing_value = None
            
        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        layout = ""
        if show_gui:
            if password:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue,key='-VALUE-', justification='c',password_char='*')],
                    [sg.Text('This field is mandatory',text_color='red')],
                    [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            elif not password and mandatory_field:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue,key='-VALUE-', justification='c')],
                    [sg.Text('This field is mandatory',text_color='red')],
                    [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            elif not password and not mandatory_field:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue,key='-VALUE-', justification='c')],
                    [sg.Submit('Done',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Close',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=True,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)

            while True:
                
                event, values = window.read()
                if event == sg.WIN_CLOSED or event == 'Close':
                    
                    if oldValue or (values and values['-VALUE-']):
                        break

                    else:
                        if mandatory_field:
                            message_pop_up("Its a mandatory field !.. Cannot proceed, exiting now..")
                            print("Exiting ClointFusion, as Mandatory field is missing")
                            sys.exit(0)
                        else:
                            print("Mandatory field is missing, continuing with None/Empty value")
                            break
                
                if event == 'Done':
                    if values['-VALUE-']:
                        break
                    else:
                        if mandatory_field:
                            message_pop_up("This value is required. Please enter the value..")
                        else:
                            break
            
            window.close()
            values['-KEY-'] = msgForUser
            
            update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-VALUE-']).strip())

            return str(values['-VALUE-']).strip()
        
        else:
            return str(existing_value)

    except Exception as ex:
        print("Error in gui_get_any_input_from_user="+str(ex))


def extract_filename_from_filepath(strFilePath=""):
    """
    Function which extracts file name from the given filepath
    """
    if strFilePath:
        try:
            strFileName = strFilePath[strFilePath.rindex("\\") + 1 : ]
            strFileName = strFileName.split(".")[0]
            return strFileName
        except Exception as ex:
            strFileName = strFilePath[strFilePath.rindex("/") + 1 : ]
            strFileName = strFileName.split(".")[0]
            return strFileName


    else:
        print("Please enter the value="+str(strFilePath))    

    

    

    
# @background
def excel_create_excel_file_in_given_folder(fullPathToTheFolder="",excelFileName="",sheet_name="Sheet1"):
    """
    Creates an excel file in the desired folder with desired filename

    Internally this uses folder_create() method to create folders if the folder/s does not exist.

    Parameters:
        fullPathToTheFolder (str) : Complete path to the folder with double slashes.
        excelFileName       (str) : File Name of the excel to be created (.xlsx extension will be added automatically.
        sheet_name           (str) : By default it will be "Sheet1".
    
    Returns:
        returns boolean TRUE if the excel file is created
    """
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title =sheet_name

        if not fullPathToTheFolder:
            fullPathToTheFolder = gui_get_folder_path_from_user('the folder to create excel file')
            
        if not excelFileName:
            excelFileName = gui_get_any_input_from_user("excel file name (without extension)")
        
        folder_create(fullPathToTheFolder)

        if ".xlsx" in excelFileName:
            excel_path = os.path.join(fullPathToTheFolder,excelFileName)
        else:
            excel_path = os.path.join(fullPathToTheFolder,excelFileName + ".xlsx")
        
        # print("Excel path="+str(excel_path))

        wb.save(filename = excel_path)
        
        return True
    except Exception as ex:
        print("Error in excel_create_excel_file_in_given_folder="+str(ex))

    
def folder_create_text_file(textFolderPath="",txtFileName=""):
    """
    Creates Text file in the given path.
    Internally this uses folder_create() method to create folders if the folder/s does not exist.
    automatically adds txt extension if not given in textFilePath.

    Parameters:
        textFilePath (str) : Complete path to the folder with double slashes.
    """
    try:

        if not textFolderPath:
            textFolderPath = gui_get_folder_path_from_user('the folder to create text file')
        
        if not txtFileName:
            txtFileName = gui_get_any_input_from_user("text file name")
            txtFileName = txtFileName 

        if ".txt" not in txtFileName:
            txtFileName = txtFileName + ".txt"
            
        f = open(os.path.join(textFolderPath, txtFileName), 'w')
        f.close()
        print("Text file created")

    except Exception as ex:
        print("Error in folder_create_text_file="+str(ex))

def _get_image_from_base64(imgFileName,imgBase64Str):
    """
    Internal function  which converts the given Base64 string to an image and saves in given path

    Parameters:
        imgFileName  (str) : image file name with png extension and optional path.
        imgBase64Str (str) : Base64 string for conversion.
    """
    if not os.path.exists(imgFileName) :
        try:
            img_binary = base64.decodebytes(imgBase64Str)
            with open(imgFileName,"wb") as f:
                f.write(img_binary)
        except Exception as ex:
            print("Error in _get_image_from_base64="+str(ex))
_get_image_from_base64(Cloint_PNG_Logo_Path,cloint_logo_base64)




        
# WatchDog : Monitors the given folder for creation / modification / deletion 
class FileMonitor_Handler(watchdog.events.PatternMatchingEventHandler):
    file_path = ""
    def __init__(self):
        watchdog.events.PatternMatchingEventHandler.__init__(self, ignore_patterns = None,
                                                     ignore_directories = False, case_sensitive = True)
    
    def on_created(self, event):
        file_path = str(event.src_path).replace("/","\\")
        print("Created : {}".format(file_path))
             
    def on_deleted(self, event):
        file_path = str(event.src_path).replace("/","\\")
        print("Deleted : {}".format(file_path))

    def on_modified(self,event):
        file_path = str(event.src_path).replace("/","\\")
        print("Modified : {}".format(file_path))

def create_batch_file(application_exe_pyw_file_path=""):
    """
    Creates .bat file for the given application / exe or even .pyw BOT developed by you. This is required in Task Scheduler.
    """

    global batch_file_path
    try:
        if not application_exe_pyw_file_path:
            application_exe_pyw_file_path = gui_get_any_file_from_user('.pyw file for which .bat is to be made','pyw')

        application_name = application_exe_pyw_file_path[application_exe_pyw_file_path.rindex("\\")+1:]

        cmd = ""

        if "exe" in application_name:
            application_name = str(application_name).replace("exe","bat")
            cmd = "start \"\" " + '"' + application_exe_pyw_file_path + '" /popup\n'

        elif "pyw" in application_name: 
            application_name = str(application_name).replace("pyw","bat")
            cmd = "start \"\" " + '"' + sys.executable + '" ' + '"' + application_exe_pyw_file_path + '" /popup\n'

        batch_file_path = os.path.join(batch_file_path,application_name)
        if not os.path.exists(batch_file_path):
            
            f = open(batch_file_path, 'w')
            f.write("@ECHO OFF\n")
            f.write("timeout 5 > nul\n")
            f.write(cmd) 
            f.write("exit")    
            f.close()

    except Exception as ex:
        print("Error in create_batch_file="+str(ex))
    

def excel_create_file(fullPathToTheFile="",fileName="",sheet_name="Sheet1"):
    try:
        if not fullPathToTheFile:
            fullPathToTheFile = gui_get_any_input_from_user('folder path to create excel')

        if not fileName:
            fileName = gui_get_any_input_from_user("Excel File Name (without extension)")

        if not os.path.exists(fullPathToTheFile):
            os.makedirs(fullPathToTheFile)
        if ".xlsx" not in fileName:
            fileName = fileName + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title =sheet_name
        fileName = os.path.join(fullPathToTheFile,fileName)
        wb.save(filename = fileName)
        print("Excel file created in the path " + fullPathToTheFile)
        return True
    except Exception as ex:
        print("Error in excel_create_file="+str(ex))
    
def folder_get_all_filenames_as_list(strFolderPath="",extension='all'):
    """
    Get all the files of the given folder in a list.

    Parameters:
        strFolderPath  (str) : Location of the folder.
        extension      (str) : extention of the file. by default all the files will be listed regarless of the extension.
    
    returns:
        allFilesOfaFolderAsLst (List) : all the file names as a list.
    """
    try:
        if not strFolderPath:
            strFolderPath = gui_get_folder_path_from_user('a folder to get all its filenames')

        if extension == "all":
            allFilesOfaFolderAsLst = [ f for f in os.listdir(strFolderPath)]
        else:
            allFilesOfaFolderAsLst = [ f for f in os.listdir(strFolderPath) if f.endswith(extension) ]

        return allFilesOfaFolderAsLst
    except Exception as ex:
        print("Error in folder_get_all_filenames_as_list="+str(ex))
    
def folder_delete_all_files(fullPathOfTheFolder="",file_extension_without_dot="all"):  
    """
    Deletes all the files of the given folder

    Parameters:
        fullPathOfTheFolder  (str) : Location of the folder.
        extension            (str) : extention of the file. by default all the files will be deleted inside the given folder 
                                    regarless of the extension.
    returns:
        count (int) : number of files deleted.
    """ 
    file_extension_with_dot = ''
    try:
        if not fullPathOfTheFolder:
            fullPathOfTheFolder = gui_get_folder_path_from_user('a folder to delete all its files')

        count = 0 
        if "." not in file_extension_without_dot :
            file_extension_with_dot = "." + file_extension_without_dot

        if file_extension_with_dot.lower() == ".all":
            filelist = [ f for f in os.listdir(fullPathOfTheFolder) ]
        else:
            filelist = [ f for f in os.listdir(fullPathOfTheFolder) if f.endswith(file_extension_with_dot) ]
        print(filelist)
        for f in filelist:
            try:
                os.remove(os.path.join(fullPathOfTheFolder, f))
                count +=1 
            except:
                pass
        
        return count
    except Exception as ex:
        print("Error in folder_delete_all_files="+str(ex)) 
        return -1
    
def message_pop_up(strMsg="",delay=3):
    """
    Specified message will popup on the screen for a specified duration of time.

    Parameters:
        strMsg  (str) : message to popup.
        delay   (int) : duration of the popup.
    """
    try:
        if not strMsg:
            strMsg = gui_get_any_input_from_user("pop-up message")
        sg.popup_no_wait(strMsg,title='ClointFusion',auto_close_duration=delay, auto_close=True, keep_on_top=True,background_color="white",text_color="black")#,icon=cloint_ico_logo_base64)
    except Exception as ex:
        print("Error in message_pop_up="+str(ex))
    
def key_hit_enter():
    """
    Enter key will be pressed once.
    """
    time.sleep(0.5)
    kb.press_and_release('enter')
    time.sleep(0.5)

def message_flash(msg="",delay=3):
    """
    specified msg will popup for a specified duration of time with OK button.

    Parameters:
        msg     (str) : message to popup.
        delay   (int) : duration of the popup.
    """
    try:
        if not msg:
            msg = gui_get_any_input_from_user("flash message")

        r = Timer(int(delay), key_hit_enter)
        r.start()
        pg.alert(text=msg, title='ClointFusion', button='OK')
    except Exception as ex:
        print("ERROR in message_flash="+str(ex))

def window_show_desktop():
    """
    Minimizes all the applications and shows Desktop.
    """
    try:
        time.sleep(0.5)
        kb.press_and_release('win+d')
        time.sleep(0.5)
    except Exception as ex:
        print("Error in window_show_desktop="+str(ex))
    
def window_get_all_opened_titles():
    """
    Gives the title of all the existing (open) windows.

    Returns:
        allTitles_lst  (list) : returns all the titles of the window as list.
    """
    try:
        allTitles_lst = []
        lst = gw.getAllTitles()
        for item in lst:
            if str(item).strip() != "" and str(item).strip() not in allTitles_lst:
                allTitles_lst.append(str(item).strip())
        return allTitles_lst
    except Exception as ex:
        print("Error in window_get_all_opened_titles="+str(ex))
    
def _window_find_exact_name(windowName=""):
    """
    Gives you the exact window name you are looking for.

    Parameters:
        windowName  (str) : Name of the window to find.

    Returns:
        win (str)              : Exact window name.
        window_found (boolean) : A boolean TRUE if the window is found
    """
    win = ""
    window_found = False

    if not windowName:
        windowName = gui_get_any_input_from_user("Partial Window Name")

    try:
        lst = gw.getAllTitles()
        
        for item in lst:
            if str(item).strip():
                if str(windowName).lower() in str(item).lower():
                    win = item
                    window_found = True
                    break
        return win, window_found
    except Exception as ex:
        print("Error in _window_find_exact_name="+str(ex))
    
def window_activate_and_maximize(windowName=""):
    """
    Activates and maximizes the desired window.

    Parameters:
        windowName  (str) : Name of the window to maximize.
    """
    try:
        if not windowName:
            windowName = gui_get_any_input_from_user("window name to Activate & Maximize")

        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]
            windw.activate()
            time.sleep(2)
            windw.maximize()
            time.sleep(2)
        else:
            print("No window OPEN by name="+str(windowName))
    except Exception as ex:
        print("Error in window_activate_and_maximize="+str(ex))
    
def window_minimize(windowName=""):
    """
    Activates and minimizes the desired window.

    Parameters:
        windowName  (str) : Name of the window to miniimize.
    """
    try:
        if not windowName:
            windowName = gui_get_any_input_from_user("window name to Minimize")
            
        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]
            windw.minimize()
            time.sleep(1)
        else:
            print("No window available to minimize by name="+str(windowName))
    except Exception as ex:
        print("Error in window_minimize="+str(ex))
    
def window_close(windowName=""):
    """
    Close the desired window.

    Parameters:
        windowName  (str) : Name of the window to close.
    """
    try:
        if not windowName:
            windowName = gui_get_any_input_from_user("window name to Close")

        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]
            windw.close()
            time.sleep(1)
        else:
            print("No window available to close, by name="+str(windowName))
    except Exception as ex:
        print("Error in window_close="+str(ex))

def launch_any_exe_bat_application(pathOfExeFile=""):
    """
    Launches any exe or batch file.

    Parameters:
        pathOfExeFile  (str) : location of the file with extension.
    """
    try:
        if not pathOfExeFile:
            pathOfExeFile = gui_get_any_file_from_user('EXE or BAT file')

        subprocess.Popen(pathOfExeFile)
        time.sleep(1) 
        window_activate_and_maximize(pathOfExeFile)
        time.sleep(1) 
    except Exception as ex:
        print("ERROR in launch_any_exe_bat_application="+str(ex))
    
class myThread1 (threading.Thread):
    def __init__(self,err_str):
        threading.Thread.__init__(self)
        self.err_str = err_str

    def run(self):
        message_flash(self.err_str)

class myThread2 (threading.Thread):
    def __init__(self,strFilePath):
        threading.Thread.__init__(self)
        self.strFilePath = strFilePath

    def run(self):
        time.sleep(1)
        img = pg.screenshot()
        time.sleep(1)

        dt_tm= str(datetime.datetime.now())    
    
        dt_tm = dt_tm.replace(" ","_")
        dt_tm = dt_tm.replace(":","-")
        dt_tm = dt_tm.split(".")[0]
        filePath = self.strFilePath + str(dt_tm)  + ".PNG"

        img.save(str(filePath))
        
def take_error_screenshot(err_str):
    """
    Takes screenshot of an error popup parallely without waiting for the flow of the program.
    The screenshot will be saved in the log folder for reference.

    Parameters:
        err_str  (str) : exception.
    """
    global error_screen_shots_path
    try:
        thread1 = myThread1(err_str)
        thread2 = myThread2(error_screen_shots_path)

        thread1.start()
        thread2.start()

        thread1.join()
        thread2.join()
    except Exception as ex:
        print("Error in take_error_screenshot="+str(ex))
    
def update_log_excel_file(message=""):
    """
    Given message will be updated in the excel log file.

    Parameters:
        message  (str) : message to update.

    Retursn:
        returns a boolean true if updated sucessfully
    """
    global fullPathToStatusLogFile
    try:
        if not message:
            message = gui_get_any_input_from_user("message to Update Log file")

        df = pd.DataFrame({'Timestamp': [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")], 'Status':[message]})
        writer = pd.ExcelWriter(fullPathToStatusLogFile, engine='openpyxl')
        writer.book = load_workbook(fullPathToStatusLogFile)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
        reader = pd.read_excel(fullPathToStatusLogFile)
        df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
        writer.close()

        print("Message updated at the row ")
        return True
    except Exception as ex:
        print("Error in update_log_excel_file="+str(ex))
        return False
    
def string_extract_only_alphabets(inputString=""):
    """
    Returns only alphabets from given input string
    """
    if not inputString:
        inputString = gui_get_any_input_from_user("input String")

    outputStr = ''.join(e for e in inputString if e.isalpha())
    return outputStr 

def string_extract_only_numbers(inputString=""):
    """
    Returns only numbers from given input string
    """
    if not inputString:
        inputString = gui_get_any_input_from_user("input String")

    outputStr = ''.join(e for e in inputString if e.isnumeric())
    return outputStr       
@lru_cache(None)
def call_otsu_threshold(img_title, is_reduce_noise=False):
    """
    OpenCV internal function for OCR
    """
    
    image = cv2.imread(img_title, 0)

    
    if is_reduce_noise:
        image = cv2.GaussianBlur(image, (5, 5), 0)

    
    _ , image_result = cv2.threshold(
        image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU,
    )
    
    cv2.imwrite(img_title, image_result)
    cv2.destroyAllWindows()

@lru_cache(None)
def read_image_cv2(img_path):
    """
    Saves the image in cv2 format.

    Parameters:
        img_path  (str) : location of the image.
    
    returns:
        image (cv2) : image in cv2 format will be returned.
    """
    if img_path and os.path.exists(img_path):
        try:
            image = cv2.imread(img_path)
            return image
        except Exception as ex:
            print("read_image_cv2 = "+str(ex))
        
    else:
        print("File not found="+str(img_path))

def ocr_get_coordinates(img_path=""):
    """
    Gets the coordinates for performing OCR on that specific region
    """
    try:
        ocr_path = os.path.join(img_folder_path,'OCR_Screenshot_1.png')

        if not img_path:
            message_counter_down_timer(start_value=5)

            img = pg.screenshot()
            img.save(ocr_path)
            img_path = ocr_path        
        if img_path and os.path.exists(img_path):
            frame=read_image_cv2(img_path)

            cv2.namedWindow('ClointFusion', cv2.WINDOW_FREERATIO) 
            cv2.setWindowProperty('ClointFusion', cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
            cv2.imshow("ClointFusion", frame)

            while True:
                key=cv2.waitKey(1) & 0xFF
                if key == ord("s"):
                    initBB = cv2.selectROI("ClointFusion", frame, fromCenter=False,showCrosshair=True)
                    print(initBB)
                    
                elif key == ord("q"):
                    break
        else:
            print("File not found="+str(img_path))

    except Exception as ex:
        print("Error in ocr_get_coordinates="+str(ex))
    
# @timeit
def ocr_magic(img_path="",x=0,y=0,w=0,h=0): #CROP specific part
    """
    Coverts the given co-ordinates /bounds of an image to OCR text. Capture the bounds using ocr_get_coordinates 

    Parameters:
        img_path  (str) : Location of the image.
        x,y,w,h   (int) : the bounds value of the croped area for ocr.
    
    Returns:
        data      (str) : the OCR processed string.
    """
    try:

        x_y_w_h = ""

        if x == 0 and y==0:
            x_y_w_h = gui_get_any_input_from_user('OCR coordinates in this format (with comma) X,Y,W,H')
            
            if x_y_w_h:
                x,y,w,h = str(x_y_w_h).split(",")
                x = int(x)
                y = int(y)
                w = int(w)
                h = int(h)

        r = (x, y, w, h)

        ocr_path = os.path.join(img_folder_path,'OCR_Screenshot.png')

        if not img_path:
            # message_counter_down_timer(start_value=3)

            img = pg.screenshot()
            img.save(ocr_path)
            img_path =ocr_path
        

        image = read_image_cv2(img_path)
        img_cropped = image[int(r[1]):int(r[1]+r[3]), int(r[0]):int(r[0]+r[2])]

        crop_path = str(x) + "_" + str(y) + "_" + str(w) + "_" + str(h)
        img_path_cropped = str(img_path).replace(".","_" + crop_path + ".")
        
        cv2.imwrite(img_path_cropped, img_cropped)

        call_otsu_threshold(img_path_cropped)

        image = read_image_cv2(img_path_cropped)
        
        data = pytesseract.image_to_string(image, lang='eng',config='--oem 3 --psm 7')
        return data
    except Exception as ex:
        print("Error in ocr_magic="+str(ex))
    
def excel_get_row_column_count(excel_path="", sheet_name="Sheet1", header=0):
    """
    Gets the row and coloumn count of the provided excel sheet.

    Parameters:
        excel_path  (str) : Full path to the excel file with slashes.
        sheet_name           (str) : by default it is Sheet1.

    Returns:
        row (int) : number of rows
        col (int) : number of coloumns
    """
    try:
        if not excel_path:
                excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user("to get row/column count")
            
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header)
        row, col = df.shape
        row = row + 1
        return row, col
    except Exception as ex:
        print("Error in excel_get_row_column_count="+str(ex))
    
def excel_copy_range_from_sheet(excel_path="",*, sheet_name='Sheet1', startCol=1, startRow=1, endCol=1, endRow=1):
    """
    Copies the specific range from the provided excel sheet and returns copied data as a list
    Parameters:
        excel_path :"Full path of the excel file with double slashes"
        sheet_name     :"Source sheet name from where contents are to be copied"
        startCol          :"Starting column number (index starts from 1) from where copying starts"
        startRow          :"Starting row number (index starts from 1) from where copying starts"
        endCol            :"Ending column number ex:4 upto where cells to be copied"
        endRow            :"Ending column number ex:5 upto where cells to be copied"

    Returns:
    rangeSelected        : the copied range data
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to copy range from')
            
        from_wb = load_workbook(filename = excel_path)
        try:
            fromSheet = from_wb[sheet_name]
        except:
            fromSheet = from_wb.worksheets[0]
        rangeSelected = []

        if endRow < startRow:
            endRow = startRow

        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(fromSheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected
    except Exception as ex:
        print("Error in copy_range_from_excel_sheet="+str(ex))
    
def excel_paste_range_to_sheet(excel_path="",*, sheet_name='Sheet1', startCol=1, startRow=1, endCol=1, endRow=1, copiedData):
    """
    Pastes the copied data in specific range of the given excel sheet.
    """
    try:
        try:
            if not excel_path:
                excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to paste range into')
                
            to_wb = load_workbook(filename = excel_path)
            toSheet = to_wb[sheet_name]
        except:
            excel_create_excel_file_in_given_folder((str(excel_path[:(str(excel_path).rindex("\\"))])),(str(excel_path[str(excel_path).rindex("\\")+1:excel_path.find(".")])),sheet_name)
            to_wb = load_workbook(filename = excel_path)
            toSheet = to_wb[sheet_name]

        if endRow < startRow:
            endRow = startRow

        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                
                toSheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        to_wb.save(excel_path)
        return countRow-1
    except Exception as ex:
        print("Error in excel_paste_range_to_sheet="+str(ex))
    
def _excel_copy_range(startCol=1, startRow=1, endCol=1, endRow=1, sheet='Sheet1'):
    """
    Copies the specific range from the given excel sheet.
    """
    try:
        rangeSelected = []
        #Loops through selected Rows
        for k in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for l in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = k, column = l).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    except Exception as ex:
        print("Error in _excel_copy_range="+str(ex))
    
def _excel_paste_range(startCol=1, startRow=1, endCol=1, endRow=1, sheetReceiving='Sheet1',copiedData=[]):
    """
    Pastes the specific range to the given excel sheet.
    """
    try:
        countRow = 0
        for k in range(startRow,endRow+1,1):
            countCol = 0
            for l in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = k, column = l).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        return countRow

    except Exception as ex:
        print("Error in _excel_paste_range="+str(ex))

def excel_split_by_column(excel_path="",*,sheet_name='Sheet1',header=0,columnName=""):
    """
    Splits the excel file by Column Name
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to split by column')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('this list of Columns (to split)',col_lst)

        data_df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,dtype=str)
        
        grouped_df = data_df.groupby(columnName)
        
        for data in  grouped_df:  
            grouped_df.get_group(data[0]).to_excel(os.path.join(output_folder_path,str(data[0]) + ".xlsx"), index=False)

    except Exception as ex:
        print("Error in excel_split_by_column="+str(ex))
    
def excel_split_the_file_on_row_count(excel_path="",*, sheet_name = 'Sheet1', rowSplitLimit="", outputFolderPath="", outputTemplateFileName ="Split"):
    """
    Splits the excel file as per given row limit
    """
    try:
        if not excel_path:
            excel_path, sheet_name, _ = gui_get_excel_sheet_header_from_user('to split on row count')
            
        if not rowSplitLimit:
            rowSplitLimit = int(gui_get_any_input_from_user("row split Count/Limit Ex: 20"))

        if not outputFolderPath:
            outputFolderPath = gui_get_folder_path_from_user('output folder to Save split excel files')

        src_wb = op.load_workbook(excel_path)
        src_ws = src_wb.worksheets[0] 

        src_ws_max_rows = src_ws.max_row
        src_ws_max_cols= src_ws.max_column 

        i = 1
        start_row = 2

        while start_row <= src_ws_max_rows:
            
            dest_wb = Workbook()
            dest_ws = dest_wb.active
            dest_ws.title = sheet_name

            #Copy ROW-1 (Header) from SOURCE to Each DESTINATION file
            selectedRange = _excel_copy_range(1,1,src_ws_max_cols,1,src_ws) #startCol, startRow, endCol, endRow, sheet
            _ =_excel_paste_range(1,1,src_ws_max_cols,1,dest_ws,selectedRange) #startCol, startRow, endCol, endRow, sheetReceiving,copiedData
            
            selectedRange = ""
            selectedRange = _excel_copy_range(1,start_row,src_ws_max_cols,start_row + rowSplitLimit - 1,src_ws) #startCol, startRow, endCol, endRow, sheet   
            _ =_excel_paste_range(1,2,src_ws_max_cols,rowSplitLimit + 1,dest_ws,selectedRange) #startCol, startRow, endCol, endRow, sheetReceiving,copiedData

            start_row = start_row + rowSplitLimit
            dest_file_name = outputFolderPath + "\\" + outputTemplateFileName + "-" + str(i) + ".xlsx"
            dest_wb.save(dest_file_name)
            print("Created " +  dest_file_name)
            i = i + 1
        return True
    except Exception as ex:
        print("Error in excel_split_the_file_on_row_count="+str(ex))
    
def excel_merge_all_files(fullPathOfTheFolder="",outputFolderPath=""):
    """
    Merges all the excel files in the given folder
    """
    try:
        if not fullPathOfTheFolder:
            fullPathOfTheFolder = gui_get_folder_path_from_user('input folder to MERGE files from')

        if not outputFolderPath:
            outputFolderPath = gui_get_folder_path_from_user('output folder to store Final merged file')
        
        filelist = [ f for f in os.listdir(fullPathOfTheFolder) if f.endswith(".xlsx") ]
        all_excel_file_lst = []
        for file1 in filelist:
            file_path = os.path.join(fullPathOfTheFolder,file1)
            print(file_path)
            all_excel_file = pd.read_excel(file_path,dtype=str)
            all_excel_file_lst.append(all_excel_file)

        appended_df = pd.concat(all_excel_file_lst)
        time_stamp_now=datetime.datetime.now().strftime("%m-%d-%Y")
        final_path= os.path.join(outputFolderPath, "Final-" + time_stamp_now + ".xlsx")
        appended_df.to_excel(final_path, index=False)
        
        return True
    except Exception as ex:
        print("Error in excel_merge_all_files="+str(ex))

def excel_drop_columns(excel_path="", sheet_name='Sheet1', header=0, columnsToBeDropped = ""):
    """
    Drops the desired column from the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('input excel to Drop the columns from')

        if not columnsToBeDropped:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnsToBeDropped = gui_get_dropdownlist_values_from_user('columns list to drop',col_lst) 

        df=pd.read_excel(excel_path,sheet_name=sheet_name, header=header)

        if isinstance(columnsToBeDropped, list):
            df.drop(columnsToBeDropped, axis = 1, inplace = True) 
        else:
            df.drop([columnsToBeDropped], axis = 1, inplace = True) 

        df.to_excel(excel_path,index=False)
    except Exception as ex:
        print("Error in excel_drop_columns="+str(ex))

def excel_sort_columns(excel_path="",*,sheet_name='Sheet1',header=0,firstColumnToBeSorted=None,secondColumnToBeSorted=None,thirdColumnToBeSorted=None,firstColumnSortType=True,secondColumnSortType=True,thirdColumnSortType=True):
    """
    A function which takes excel full path to excel and column names on which sort is to be performed

    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to sort the column')

        if not firstColumnToBeSorted:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            usecols = gui_get_dropdownlist_values_from_user('minimum 1 and maximum 3 columns to sort',col_lst)
            
            if len(usecols) == 3:
                firstColumnToBeSorted , secondColumnToBeSorted , thirdColumnToBeSorted = usecols
            elif len(usecols) == 2:
                firstColumnToBeSorted , secondColumnToBeSorted = usecols
            elif len(usecols) == 1:
                firstColumnToBeSorted = usecols[0]
        df=pd.read_excel(excel_path,sheet_name=sheet_name, header=header)
        if thirdColumnToBeSorted is not None and secondColumnToBeSorted is not None and firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted,secondColumnToBeSorted,thirdColumnToBeSorted],ascending=[firstColumnSortType,secondColumnSortType,thirdColumnSortType])
        
        elif secondColumnToBeSorted is not None and firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted,secondColumnToBeSorted],ascending=[firstColumnSortType,secondColumnSortType])
        
        elif firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted],ascending=[firstColumnSortType])

        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = load_workbook(excel_path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
        df.to_excel(writer,sheet_name=sheet_name,index=False)

        writer.save()
        writer.close()    
        return True
    except Exception as ex:
        print("Error in excel_sort_columns="+str(ex))        
    
def excel_clear_sheet(excel_path="",sheet_name="Sheet1", header=0):
    """
    Clears the contents of given excel files keeping header row intact
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to clear the sheet')

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header) 
        df = df.head(0)

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        # writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        # writer.book = load_workbook(excel_path)
        # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # df.to_excel(writer,sheet_name=sheet_name,index=False)

        # writer.save()
        # writer.close()

    except Exception as ex:
        print("Error in excel_clear_sheet="+str(ex))

def excel_set_single_cell(excel_path="", *, sheet_name="Sheet1", header=0, columnName="", cellNumber=0, setText=""): 
    """
    Writes the given text to the desired column/cell number for the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to set cell')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to set vlaue',col_lst,multi_select=False)   

        if not setText:
            setText = gui_get_any_input_from_user("text value to set the cell")

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header)
        
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer.book = load_workbook(excel_path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        
        df.at[cellNumber,columnName] = setText
        df.to_excel(writer, sheet_name=sheet_name ,index=False)    
        writer.save()
        writer.close()
        return True

    except Exception as ex:
        print("Error in excel_set_single_cell="+str(ex))

def excel_get_single_cell(excel_path="",*,sheet_name="Sheet1",header=0, columnName="",cellNumber=0): 
    """
    Gets the text from the desired column/cell number of the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to get cell')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to get vlaue',col_lst,multi_select=False)   

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,usecols={columnName[0]})
        cellValue = df.at[cellNumber,columnName[0]]
        return cellValue
    except Exception as ex:
        print("Error in excel_get_single_cell="+str(ex))

def excel_remove_duplicates(excel_path="", *, sheet_name="Sheet1", header=0, columnName="", saveResultsInSameExcel=True, which_one_to_keep="first"): 
    """
    Drops the duplicates from the desired Column of the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to remove duplicates')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to remove duplicates',col_lst)  

        df = pd.read_excel(excel_path, sheet_name=sheet_name,header=header) 
        count = 0 
        if saveResultsInSameExcel:
            df.drop_duplicates(subset=columnName, keep=which_one_to_keep, inplace=True)
            df.to_excel(excel_path, index=False)
            count = df.shape[0]
        else:
            df1 = df.drop_duplicates(subset=columnName, keep=which_one_to_keep, inplace=False)
            excel_path = str(excel_path).replace(".","_DupDropped.")
            df1.to_excel(excel_path, index=False)
            count = df1.shape[0]

        print(str(count) + " rows affected")
        return count
    except Exception as ex:
        print("Error in excel_remove_duplicates="+str(ex))
    
def excel_vlook_up(filepath_1="", *, sheet_name_1 = 'Sheet1', header_1 = 0, filepath_2="", sheet_name_2 = 'Sheet1', header_2 = 0, Output_path="", OutputExcelFileName="", match_column_name="",how='left'):
    """
    Performs excel_vlook_up on the given excel files for the desired columns. Possible values for how are "inner","left", "right", "outer"
    """
    try:
        if not filepath_1:
            filepath_1, sheet_name_1, header_1 = gui_get_excel_sheet_header_from_user('(Vlookup) first excel')
             
        if not filepath_2:
            filepath_2, sheet_name_2, header_2 = gui_get_excel_sheet_header_from_user('(Vlookup) second excel')
            
        if not match_column_name:
            col_lst = excel_get_all_header_columns(filepath_1, sheet_name_1, header_1)
            match_column_name = gui_get_dropdownlist_values_from_user('Vlookup column name to be matched',col_lst,multi_select=False) 
            match_column_name = match_column_name[0]
        df1 = pd.read_excel(filepath_1, sheet_name = sheet_name_1, header = header_1)
        df2 = pd.read_excel(filepath_2, sheet_name = sheet_name_2, header = header_2)

        df = pd.merge(df1, df2, on= match_column_name, how = how)

        output_file_path = ""
        if str(OutputExcelFileName).endswith(".*"):
            OutputExcelFileName = OutputExcelFileName.split(".")[0]
        
        if Output_path and OutputExcelFileName:
            if ".xlsx" in OutputExcelFileName:
                output_file_path = os.path.join(Output_path, OutputExcelFileName)
            else:
                output_file_path = os.path.join(Output_path, OutputExcelFileName  + ".xlsx")

        else:
            output_file_path = filepath_1

        # with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
        #     df.to_excel(writer, index=False)

        df.to_excel(output_file_path, index = False)
        print("excel_vlook_up Done")
        return True
    
    except Exception as ex:
        print("Error in excel_vlook_up="+str(ex))
    
def screen_clear_search(delay=0.2):
    """
    Clears previously found text (crtl+f highlight)
    """
    try:
        kb.press_and_release("ctrl+f")
        time.sleep(delay)
        pg.typewrite("^%#")
        time.sleep(delay)
        kb.press_and_release("esc")
        time.sleep(delay)
    except Exception as ex:
        print("Error in screen_clear_search="+str(ex))
    
def scrape_save_contents_to_notepad(folderPathToSaveTheNotepad=""): #"Full path to the folder (with double slashes) where notepad is to be stored"
    """
    Copy pastes all the available text on the screen to notepad and saves it.
    """
    try:
        if not folderPathToSaveTheNotepad:
            folderPathToSaveTheNotepad = gui_get_folder_path_from_user('folder to save notepad contents')

        message_counter_down_timer(3)
        time.sleep(1)
        screen_size=pg.size()
        pg.click(screen_size[0]/2,screen_size[1]/2)
        time.sleep(0.5)

        kb.press_and_release("ctrl+a")
        time.sleep(1)
        kb.press_and_release("ctrl+c")
        time.sleep(1)
        
        clipboard_data = clipboard.paste()
        time.sleep(2)
        
        screen_clear_search()

        if str(folderPathToSaveTheNotepad).endswith("\\"):
            notepad_file_path = folderPathToSaveTheNotepad + "notepad-contents.txt"
        else:
            notepad_file_path = folderPathToSaveTheNotepad + "\\notepad-contents.txt"

        f = open(notepad_file_path, "w", encoding="utf-8")
        f.write(clipboard_data)
        time.sleep(10)
        f.close()

        clipboard_data = ''
        return "Saved the contents at " + notepad_file_path
    except Exception as ex:
        print("Error in scrape_SaveContentsToNotepad = "+str(ex))
    
def scrape_get_contents_by_search_copy_paste(highlightText=""):
    """
    Gets the focus on the screen by searching given text using crtl+f and performs copy/paste of all data. Useful in Citrix applications
    This is useful in Citrix applications
    """
    output_lst_newline_removed = []
    try:
        if not highlightText:
            highlightText = gui_get_any_input_from_user("text to be searched in Citrix environment")

        time.sleep(1)
        kb.press_and_release("ctrl+f")
        time.sleep(1)
        pg.typewrite(highlightText)
        time.sleep(1)
        kb.press_and_release("enter")
        time.sleep(1)
        kb.press_and_release("esc")
        time.sleep(2)

        pg.PAUSE = 2
        kb.press_and_release("ctrl+a")
        time.sleep(2)
        kb.press_and_release("ctrl+c")
        time.sleep(2)
        
        clipboard_data = clipboard.paste()
        time.sleep(2)
        
        screen_clear_search()

        entire_data_as_list= clipboard_data.splitlines()
        for line in entire_data_as_list:
            if line.strip():
                output_lst_newline_removed.append(line.strip())

        clipboard_data = ''
        return output_lst_newline_removed
    except Exception as ex:
        print("Error in scrape_get_contents_by_search_copy_paste="+str(ex))
    
def mouse_move(x="",y=""):
    """
    Moves the cursor to the given X Y Co-ordinates.
    """
    try:
        if not x and not y:
            x_y = str(gui_get_any_input_from_user("X,Y co-ordinates to the move Mouse to. Ex: 200,215"))
            if "," in x_y:
                x, y = x_y.split(",")
                x = int(x)
                y = int(y)
            else:
                x = x_y.split(" ")[0]
                y = x_y.split(" ")[1]
        if x and y:
            time.sleep(0.2)
            pg.moveTo(x,y)
            time.sleep(0.2)
    except Exception as ex:
        print("Error in mouse_move="+str(ex))
    
def mouse_get_color_by_position(pos=[]):
    """
    Gets the color by X Y co-ordinates of the screen.
    """
    try:
        if not pos:
            pos1 = gui_get_any_input_from_user("X,Y co-ordinates to get its color. Ex: 200,215")
            pos = tuple(map(int, pos1.split(',')))

        im = pg.screenshot()
        time.sleep(0.5)
        return im.getpixel(pos)    
    except Exception as ex:
        print("Error in mouse_get_color_by_position = "+str(ex))
    
def mouse_click(x="", y="", left_or_right="left", single_double_triple="single", copyToClipBoard_Yes_No="no"):
    """
    Clicks at the given X Y Co-ordinates on the screen using ingle / double / tripple click(s).
    Optionally copies selected data to clipboard (works for double / triple clicks)
    """
    try:
        if not x and not y:
            x_y = str(gui_get_any_input_from_user("X,Y co-ordinates to perform Mouse (Left) Click. Ex: 200,215"))
            if "," in x_y:
                x, y = x_y.split(",")
                x = int(x)
                y = int(y)
            else:
                x = x_y.split(" ")[0]
                y = x_y.split(" ")[1]

        copiedText = ""
        time.sleep(1)

        if x and y:
            if single_double_triple.lower() == "single" and left_or_right.lower() == "left":
                pg.click(x,y)
            elif single_double_triple.lower() == "double" and left_or_right.lower() == "left":
                pg.doubleClick(x,y)
            elif single_double_triple.lower() == "triple" and left_or_right.lower() == "left":
                pg.tripleClick(x,y)
            elif single_double_triple.lower() == "single" and left_or_right.lower() == "right":
                pg.rightClick(x,y)
            time.sleep(1)    

            if copyToClipBoard_Yes_No.lower() == "yes":
                kb.press_and_release("ctrl+c")
                time.sleep(1)
                copiedText = clipboard.paste().strip()
                time.sleep(1)
                
            time.sleep(1)    
            return copiedText
    except Exception as ex:
        print("Error in mouseClick="+str(ex))
    
def mouse_drag_from_to(X1="",Y1="",X2="",Y2="",delay=0.5):
    """
    Clicks and drags from X1 Y1 co-ordinates to X2 Y2 Co-ordinates on the screen
    """
    try:
        if not X1 and not Y1:
            x_y = str(gui_get_any_input_from_user("Mouse Drag FROM Values ex: 200,215"))
            if "," in x_y:
                X1, Y1 = x_y.split(",")
                X1 = int(X1)
                Y1 = int(Y1)

        if not X2 and not Y2:
            x_y = str(gui_get_any_input_from_user("Mouse Drag TO Values ex: 200,215"))
            if "," in x_y:
                X2, Y2 = x_y.split(",")
                X2 = int(X2)
                Y2 = int(Y2)
        time.sleep(0.2)
        pg.moveTo(X1,Y1,duration=delay)
        pg.dragTo(X2,Y2,duration=delay,button='left')
        time.sleep(0.2)
    except Exception as ex:
        print("Error in mouse_drag_from_to="+str(ex))
    
def search_highlight_tab_enter_open(searchText="",hitEnterKey="Yes"):
    """
    Searches for a text on screen using crtl+f and hits enter.
    This function is useful in Citrix environment
    """
    try:
        if not searchText:
            searchText = gui_get_any_input_from_user("Search Text to Highlight (in Citrix Environment)")

        time.sleep(0.5)
        kb.press_and_release("ctrl+f")
        time.sleep(0.5)
        kb.write(searchText)
        time.sleep(0.5)
        kb.press_and_release("enter")
        time.sleep(0.5)
        kb.press_and_release("esc")
        time.sleep(0.2)
        if hitEnterKey.lower() == "yes":
            kb.press_and_release("tab")
            time.sleep(0.3)
            kb.press_and_release("shift+tab")
            time.sleep(0.3)
            kb.press_and_release("enter")
            time.sleep(2)
        return True

    except Exception as ex:
        print("Error in search_highlight_tab_enter_open="+str(ex))
    
def key_press(strKeys=""):
    """
    Emulates the given keystrokes.
    """
    try:
        if not strKeys:            
            strKeys = gui_get_any_input_from_user("keys combination using + as delimeter. Ex: ctrl+O")

        strKeys = strKeys.lower()
        if "shift" in strKeys:
            strKeys = strKeys.replace("shift","left shift+right shift")

        time.sleep(0.5)
        kb.press_and_release(strKeys)
        time.sleep(0.5)
    except Exception as ex:
        print("Error in key_press="+str(ex))
    
def key_write_enter(strMsg="",delay=1,key="e"):
    """
    Writes/Types the given text and press enter (by default) or tab key.
    """
    try:
        if not strMsg:
            strMsg = gui_get_any_input_from_user("message / username / any text")

        time.sleep(0.2)
        kb.write(strMsg)
        time.sleep(delay)
        if key.lower() == "e":
            key_press('enter')
        elif key.lower() == "t":
            key_press('tab')
        time.sleep(1)
    except Exception as ex:
        print("Error in key_write_enter="+str(ex))

def date_convert_to_US_format(input_str=""):
    """
    Converts the given date to US date format.
    """
    try:
        if not input_str:
            input_str = gui_get_any_input_from_user('Date value Ex: 01/01/2021')
        match = re.search(r'\d{4}-\d{2}-\d{2}', input_str) #1
        if match == None:
            match = re.search(r'\d{2}-\d{2}-\d{4}', input_str) #2
            if match == None:
                match = re.search(r'\d{2}/\d{2}/\d{4}', input_str) #3
                if match == None:
                    match = re.search(r'\d{4}/\d{2}/\d{2}', input_str) #4
                    if match == None:
                        match = re.findall(r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s\d\d,\s\d{4}',input_str) #5
                        dt=datetime.datetime.strptime(match[0], '%b %d, %Y').date() #5 Jan 01, 2020
                    else:    
                        dt=datetime.datetime.strptime(match.group(), '%Y/%m/%d').date() #4
                else:
                    try:
                        dt=datetime.datetime.strptime(match.group(),'%d/%m/%Y').date() #3
                    except:
                        dt=datetime.datetime.strptime(match.group(),'%m/%d/%Y').date() #3
            else:
                try:
                    dt=datetime.datetime.strptime(match.group(), '%d-%m-%Y').date()#2
                except:
                    dt=datetime.datetime.strptime(match.group(), '%m-%d-%Y').date()#2
        else:
            dt=datetime.datetime.strptime(match.group(), '%Y-%m-%d').date() #1
        return dt.strftime('%m/%d/%Y')    
    except Exception as ex:
        print("Error in date_convert_to_US_format="+str(ex))
    
def mouse_search_snip_return_coordinates_x_y(img="", conf=0.9, wait=180,region=(0,0,pg.size()[0],pg.size()[1])): #180
    """
    Searches the given image on the screen and returns its center of X Y co-ordinates.
    """
    try:
        if not img:
            img = gui_get_any_file_from_user("snip image file, to get X,Y coordinates","png")

        time.sleep(1)
        pos = pg.locateOnScreen(img,confidence=conf,region=region) 
        i = 0
        while pos == None and i < int(wait):
            pos = ()
            pos = pg.locateOnScreen(img, confidence=conf,region=region)   
            time.sleep(1)
            i = i + 1

        time.sleep(1)

        if pos:
            x,y = pos.left + int(pos.width / 2), pos.top + int(pos.height / 2)
            pos = ()
            pos=(x,y)
            
            return pos
        return pos
    except Exception as ex:
        print("Error in mouse_search_snip_return_coordinates_x_y="+str(ex))

    
def find_text_on_screen(searchText="",delay=0.1, occurance=1,isSearchToBeCleared=False):
    """
    Clears previous search and finds the provided text on screen.
    """
    screen_clear_search() #default

    if not searchText:
        searchText = gui_get_any_input_from_user("search text to Find on screen")

    time.sleep(delay)
    kb.press_and_release("ctrl+f")
    time.sleep(delay)
    pg.typewrite(searchText)
    time.sleep(delay)

    for i in range(occurance-1):
        kb.press_and_release("enter")
        time.sleep(delay)

    kb.press_and_release("esc")
    time.sleep(delay)

    if isSearchToBeCleared:
        screen_clear_search()

def mouse_search_snip_return_coordinates_box(img="", conf=0.9, wait=180,region=(0,0,pg.size()[0],pg.size()[1])):
    """
    Searches the given image on the screen and returns the 4 bounds co-ordinates (x,y,w,h)
    """
    try:
        if not img:
            img = gui_get_any_file_from_user("snip image file, to get BOX coordinates","png")
        time.sleep(1)
        
        pos = pg.locateOnScreen(img,confidence=conf,region=region) 
        i = 0
        while pos == None and i < int(wait):
            pos = ()
            pos = pg.locateOnScreen(img, confidence=conf,region=region)   
            time.sleep(1)
            i = i + 1
        time.sleep(1)
        return pos

    except Exception as ex:
        print("Error in mouse_search_snip_return_coordinates_box="+str(ex))

def mouse_find_highlight_click(searchText="",delay=0.1,occurance=1):
    """
    Searches the given text on the screen, highlights and clicks it.
    """  
    try:
        if not searchText:
            searchText = gui_get_any_input_from_user("search text to Highlight & Click")


        time.sleep(0.2)

        find_text_on_screen(searchText,delay=delay,occurance=occurance,isSearchToBeCleared = True) #clear the search

        img = pg.screenshot()
        img.save(ss_path_b)
        time.sleep(0.2)
        imageA = cv2.imread(ss_path_b)
        time.sleep(0.2)

        find_text_on_screen(searchText,delay=delay,occurance=occurance,isSearchToBeCleared = False) #dont clear the searched text

        img = pg.screenshot()
        img.save(ss_path_a)
        time.sleep(0.2)
        imageB = cv2.imread(ss_path_a)
        time.sleep(0.2)

        # convert both images to grayscale
        grayA = cv2.cvtColor(imageA, cv2.COLOR_BGR2GRAY)
        grayB = cv2.cvtColor(imageB, cv2.COLOR_BGR2GRAY)

        # compute the Structural Similarity Index (SSIM) between the two
        (_, diff) = structural_similarity(grayA, grayB, full=True)
        diff = (diff * 255).astype("uint8")

        thresh = cv2.threshold(diff, 0, 255,
            cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL,
            cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        # loop over the contours
        for c in cnts:
            (x, y, w, h) = cv2.boundingRect(c)
            
            X = int(x + (w/2))
            Y = int(y + (h/2))
            
            pg.click(X, Y)
            time.sleep(0.5)
            break

    except Exception as ex:
        print("Error in mouse_find_highlight_click="+str(ex))
    
#Web Browser Automation Functions
def _start_chrome_service(dp=False,dn=True,igc=True,smcp=True,i=False,h=False,whtsp=False, profile_path=""):        #pass boolean True or False
    """
    Starts the chrome driver service and launches the chrome browser with specified arguments.
    """
    global driver
    global service
    try:
        options = webdriver.ChromeOptions()
        chromedriver_autoinstaller.install()
        chrome_driver_path = which(chromedriver_autoinstaller.utils.get_chromedriver_filename())
        service = Service(chrome_driver_path)
        service.start()

        #False
        if dp:
            options.add_argument("--disable-popup-blocking")   
        #True             
        if dn:  
            options.add_argument("--disable-notifications")          
        #True       
        if igc:
            options.add_argument("--ignore-certificate-errors")    
        #True         
        if smcp:
            options.add_argument("--suppress-message-center-popups")        
        #False
        if i:
            options.add_argument("--incognito")     
        #False                        
        if h:
            options.add_argument("--headless")                              
            options.add_argument("--disable-gpu")
        
        if whtsp :
            options.add_argument("--user-data-dir={}".format(profile_path))
            options.add_argument('--profile-directory=Profile 108')        
        options.add_argument("--disable-translate")
        #True
        options.add_argument("--start-maximized")                           
        #True
        options.add_argument("--ignore-autocomplete-off-autofill")          
        #True
        options.add_argument("--no-first-run")             
        #options.add_argument("--window-size=1920,1080")
        
        driver = webdriver.Remote(service.service_url,options=options)
    except Exception as ex:
        print("Error in _start_chrome_service = "+str(ex))
        driver.quit
        

    
def browser_get_page_source_html():
    """
    Gets the complete html page source of the given page as string.
    """
    try:
        #Get whole HTML Source in the page as List
        return driver.page_source           
    except Exception as ex:
        print("Error in browser_get_page_source_html = "+str(ex))
    
def browser_get_page_source_text():
    """
    Gets complete text of the given page as list.
    """
    try:                                     
        #Get whole text in the page as List
        text_items = []
        data = driver.find_elements_by_tag_name('html')
        for item in data:
            text_items.append(item.text)
        return text_items
    except Exception as ex:
        print("Error in browser_get_page_source_text = "+str(ex))

def browser_navigate_s(nav):
    """
    Navigates to given url or goes forward, backward or refresh according to the specified argument.
    """
    try:
        if len(nav) >5 :
            #Navigate to URL
            driver.get(nav)                 
            return
        if nav.lower() == "b" :             
            #Navigate Back
            driver.back()
            return
        if nav.lower() == "f" :       
            #Navigate Forward       
            driver.forward()
            return
        if nav.lower() == "r" :    
            #Refresh Page         
            driver.refresh()
            return
    except Exception as ex:
        print("Error in browser_navigate_s = "+str(ex))
        driver.quit()
        Chrome_Service_Started = False

#returns a single element    
def browser_locate_element_s(element="",type_="xpath"):   
    """
    Locates the XPATH element on screen and returns the element.
    """
    try:                                                 
        if not element:
            element = gui_get_any_input_from_user('browser Element to locate (Selenium)')

        if type_.lower() == "xpath":                                        
            #find by Xpath    
            element1 = driver.find_element_by_xpath(element)                
            return element1 
        if type_.lower() == "href":
            element1 = driver.find_element_by_link_text(element)  
            #find by link text href          
            return element1
        if type_.lower() == "css":               
            #find by css selector                           
            element1 = driver.find_element_by_css_selector(element)
            return element1
    except Exception as ex:
        print("Error in browser_locate_element_s = "+str(ex))

#returns multiple elements as a list
def browser_locate_elements_s(element="",type_="xpath"):                            
    """
    Locates the XPATH elements on screen and return the elements.
    """
    try:
        if not element:
            element = gui_get_any_input_from_user('browser ElementS to locate (Selenium)')

        if type_.lower() == "xpath":
            elements = driver.find_elements_by_xpath(element)               
            return elements
        if type_.lower() == "href":
            #find by link text href
            elements = driver.find_element_by_link_text(element)            
            return elements
        if type_.lower() == "css":          
            #find by css selector                                
            elements = driver.find_element_by_css_selector(element)
            return elements
    except Exception as ex:
        print("Error in browser_locate_elements_s = "+str(ex))

#Click or Send enter key to the element
def browser_mouse_click_s(element="",i="c"):   
    """
    Clicks on given web element using XPath
    """
    try:                                              
        if not element:
            element = gui_get_any_input_from_user('browser element for Mouse Click')
        if i == "c":
            element.click()                                                 
            return
        if i == "e":
            element.send_keys(Keys.ENTER)                                   
            return
    except Exception as ex:
        print("Error in browser_mouse_click_s = "+str(ex))
    
def browser_write_s(element="",write=""):
    try:
        if not element:
            element = gui_get_any_input_from_user('browser Element to Write to')

        if not write:
            write = gui_get_any_input_from_user('Text value to write to browser element')

        #write text in text fields
        element.send_keys(write)                
        return
    except Exception as ex:
        print("Error in browser_write_s = "+str(ex))
    
def browser_wait_s():
    try:
        driver.implicitly_wait(120)
        return
    except Exception as ex:
        print("Error in browser_wait_s = "+str(ex))
    
def launch_website_s(URL1="",dp=False,dn=True,igc=True,smcp=True,i=False,h=False,whtsp=False, profile_path=""):
    """
    Starts the chrome service, opens the browser and launches your website URL. This is the first function to be called for any web-related BOT Development
    """
    global Chrome_Service_Started

    try:
        if not URL1:
            URL1 = gui_get_any_input_from_user("website URL to Launch Website using Selenium. Ex https://www.google.com")

        if not Chrome_Service_Started: 
            _start_chrome_service(dp=dp,dn=dn,igc=igc,smcp=smcp,i=i,h=h,whtsp=whtsp, profile_path=profile_path)
            browser_wait_s()
            Chrome_Service_Started = True

        browser_navigate_s(URL1)
        
        browser_wait_s()
        
    except Exception as ex:
        print("Error in launch_website_s="+str(ex))
        driver.quit()
        Chrome_Service_Started = False
    
def _search_image(img,confidence):  
    """
    Internal Function
    """
    try:
        w,h = pg.size()  
        im = region_grabber((0, 0, w, h))
        pos = imagesearcharea(img, 0, 0, w, h, confidence, im)
        if pos[0] > 0 and pos[1] > 0 :
            return pos
    except Exception as ex:
        print("Errror in _search_image="+str(ex))
    
#searches multiple images and returns list of tuple for all images found
def search_multiple_images_in_parallel(img_lst=[], confidence=0.9):
    """
    Returns the postion of all the images passed as list
    """
    try:
        if not img_lst:
            img_lst = []
            img_lst_folder_path = gui_get_folder_path_from_user('folder containing image(s) to search')
            img_files_lst = folder_get_all_filenames_as_list(img_lst_folder_path,".png")
            for img_file in img_files_lst:
                img_file = os.path.join(img_lst_folder_path,img_file)
                img_file = img_file.replace("/","\\")
                img_file = img_file.replace("\\","\\")
                img_lst.append(img_file)
            time.sleep(1)
        if len(img_lst) > 0 :
            results = Parallel(n_jobs=10)(delayed(_search_image)(img,confidence) for img in img_lst)
            return results
    except Exception as ex:
        print("Errror in search_multiple_images_in_parallel="+str(ex))
    
def _locate_image(snip_url, ai_screenshot,confidence):
    """
    Internal function
    """
    try:
        cordinates = pg.locate(snip_url, ai_screenshot, confidence=confidence)
        if cordinates is None:
            return None
        return cordinates
    except Exception as ex:
        print("Error in _locate_image="+str(ex))
    
def _predict_ai_coordinates():
    """
    Internal function
    """
    try:
        x,y,c,m,n = 0,0,0,0,0
        _,_,w,h = 0,0,0,0
        global ai_processes
        for task in as_completed(ai_processes):
            if task.result() is not None:
                if m == 0:
                    m,n = pg.center(task.result())
                    l,t,w,h = task.result()
                a,b = pg.center(task.result())
                if ((m+(w//4)+1) >= a and (m-(w//4)-1) <= a) and ((n+(h//4)+1) >= b and (n-(h//4)-1) <= b):
                    pass
                else:
                    return "Multiple images detected, use mouse_ai_locate_multiple_images_on_screen() function","Multiple images detected, use mouse_ai_locate_multiple_images_on_screen() function"
                c += 1
                x += a
                y += b
        if (x>0) and (y>0):
            ai_x = x//c
            ai_y = y//c
            return ai_x,ai_y
        else:
            return None,None

    except Exception as ex:
        print("Error in _predict_ai_coordinates="+str(ex))
    
def ai_screenshot():
    try:
        global ai_screenshot
        ai_screenshot = pg.screenshot()
        return
    except Exception as ex:
        print("Error in ai_screenshot="+str(ex))
    
def _multitreading_locateimage(ai_snip_list,confidence=.8):    
    """
    internal function
    """
    try:
        with ThreadPoolExecutor(max_workers=50) as executor:
            for snip in ai_snip_list:
                snip_url = snip
                ai_processes.append(executor.submit(_locate_image,snip_url,ai_screenshot,confidence=confidence))
    except Exception as ex:
        print("Error in _multitreading_locateimage="+str(ex))
    
def mouse_ai_locate_snip_on_screen(ai_snip_list=[], confidence=.8):
    try:
        if not ai_snip_list:
            ai_snip_list = []
            img_lst_folder_path = gui_get_folder_path_from_user('folder containing similar image(s) with minor variations')

            # img_lst_folder_path = gui_get_any_file_from_user('snip to locate on screen')

            img_files_lst = folder_get_all_filenames_as_list(img_lst_folder_path,".png")
            for img_file in img_files_lst:
                img_file = os.path.join(img_lst_folder_path,img_file)
                img_file = img_file.replace("/","\\")
                img_file = img_file.replace("\\","\\")
                ai_snip_list.append(img_file)
            time.sleep(0.5)

        ai_screenshot()
        _multitreading_locateimage(ai_snip_list=ai_snip_list,confidence=confidence)
        ai_x,ai_y = _predict_ai_coordinates() 
        return ai_x,ai_y
    except Exception as ex:
        print("Error in mouse_ai_locate_snip_on_screen="+str(ex))
    
def mouse_ai_locate_multiple_images_on_screen(ai_snip_list=[],confidence=.8,click=False):
    try:
        ai_multiple_x_y = []
        if not ai_snip_list:
            ai_snip_list = []
            img_lst_folder_path = gui_get_folder_path_from_user('folder containing different image(s) to locate on screen')
            img_files_lst = folder_get_all_filenames_as_list(img_lst_folder_path,".png")
            for img_file in img_files_lst:
                img_file = os.path.join(img_lst_folder_path,img_file)
                img_file = img_file.replace("/","\\")
                img_file = img_file.replace("\\","\\")
                ai_snip_list.append(img_file)
            time.sleep(1)
        ai_screenshot()
        _multitreading_locateimage(ai_snip_list=ai_snip_list,confidence=confidence)
        for task in as_completed(ai_processes):
            if task.result() is not None:
                ai_multiple_x_y.append(pg.center(task.result()))
        if ai_multiple_x_y == []:
            ai_multiple_x_y = None
        # print(ai_multiple_x_y)
        if ai_multiple_x_y and click==True:
            for click in ai_multiple_x_y:
                time.sleep(0.5)
                pg.click(click)
        return ai_multiple_x_y
    except Exception as ex:
        print("Error in mouse_ai_locate_multiple_images_on_screen="+str(ex))
    
def schedule_create_task(Weekly_Daily="D",*,week_day="Sun",start_time_hh_mm_24_hr_frmt="11:00"):
    """
    Schedules (weekly & daily options as of now) the current BOT (.bat) using Windows Task Scheduler. Please call create_batch_file() function before using this function to convert .pyw file to .bat
    """
    global batch_file_path
    try:


        str_cmd = ""

        if not batch_file_path:
            batch_file_path = gui_get_any_file_from_user('BATCH file to Schedule. Please call create_batch_file() to create one')

        if Weekly_Daily == "D":
            str_cmd = r"powershell.exe Start-Process schtasks '/create  /SC DAILY /tn ClointFusion\{} /tr {} /st {}' ".format(bot_name,batch_file_path,start_time_hh_mm_24_hr_frmt)
        elif Weekly_Daily == "W":
            str_cmd = r"powershell.exe Start-Process schtasks '/create  /SC WEEKLY /D {} /tn ClointFusion\{} /tr {} /st {}' ".format(week_day,bot_name,batch_file_path,start_time_hh_mm_24_hr_frmt)

        subprocess.call(str_cmd)
        print("Task Scheduled")
    except Exception as ex:
        print("Error in schedule_create_task="+str(ex))

def schedule_delete_task():
    """
    Deletes already scheduled task. Asks user to supply task_name used during scheduling the task. You can also perform this action from Windows Task Scheduler.
    """
    try:
        str_cmd = r"powershell.exe Start-Process schtasks '/delete /tn ClointFusion\{} ' ".format(bot_name)
        
        subprocess.call(str_cmd)
        print("Task {} Deleted".format(bot_name))

    except Exception as ex:
        print("Error in schedule_delete_task="+str(ex))

@lru_cache(None)
def _get_tabular_data_from_website(Website_URL):
    """
    internal function
    """
    all_tables = ""
    try:
        all_tables = pd.read_html(Website_URL)
        return all_tables
    except Exception as ex:
        print("Error in _get_tabular_data_from_website="+str(ex))

def browser_get_html_tabular_data_from_website(Website_URL="",table_index=-1,drop_first_row=False,drop_first_few_rows=[0],drop_last_row=False):
    """
    Web Scrape HTML Tables : Gets Website Table Data Easily as an Excel using Pandas. Just pass the URL of Website having HTML Tables.
    If there are 5 tables on that HTML page and you want 4th table, pass table_index as 3

    Ex: browser_get_html_tabular_data_from_website(Website_URL=URL)
    """
    try:
        if not Website_URL:            
            Website_URL= gui_get_any_input_from_user("website URL to get HTML Tabular Data ex: https://www.google.com ")

        all_tables = _get_tabular_data_from_website(Website_URL)

        if all_tables:
            
            # if no table_index is specified, then get all tables in output
            if table_index == -1:
                strFileName = Website_URL[Website_URL.rindex("/")+1:] + "_All_Tables" +  ".xlsx"
                excel_create_excel_file_in_given_folder(output_folder_path,strFileName)
            else:
                strFileName = Website_URL[Website_URL.rindex("/")+1:] + "_" + str(table_index) +  ".xlsx"

            strFileName = os.path.join(output_folder_path,strFileName)
            
            if table_index == -1:
                for i in range(len(all_tables)):
                    table = all_tables[i] #lool thru table_index values

                    table = table.reset_index(drop=True) #Avoid multi index error in our dataframes

                    with pd.ExcelWriter(strFileName, engine='openpyxl', mode='a') as writer:
                        table.to_excel(writer, sheet_name=str(i)) #index=False
            else:
                table = all_tables[table_index] #get required table_index
                
                if drop_first_row:
                    table = table.drop(drop_first_few_rows) # Drop first few rows (passed as list)

                if drop_last_row:
                    table = table.drop(len(table)-1) # Drop last row

            # table.columns = list(table.iloc[0])
            # table = table.drop(len(drop_first_few_rows)) 

                table = table.reset_index(drop=True) 

                table.to_excel(strFileName, index=False)

            print("Table saved as Excel at {} ".format(strFileName))

        else:
            print("No tables found in given website " + str(Website_URL))

    except Exception as ex:
        print("Error in browser_get_html_tabular_data_from_website="+str(ex))

def message_send_whatsapp(name="",message=""):
    """
    SendWhatsAppMessage function to send message from WhatsApp web application
    """
    try:
        if not name:
            name = gui_get_any_input_from_user('Exact contact Name to send your message')

        if not message:
            message= gui_get_any_input_from_user('your Message')

        path = os.environ['USERPROFILE']
        cprofile_path = os.path.join(path,"AppData","Local","Google","Chrome","User Data","Default2")

        contact_xpath = "//div//span[@title='{}']".format(name)
        message_xpath = "//div[@dir='ltr'][@contenteditable='true'][@spellcheck='true']"

        launch_website_s(URL1="https://web.whatsapp.com/",whtsp=True,profile_path=cprofile_path)
        browser_navigate_s("https://web.whatsapp.com/")
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, contact_xpath)))
        contact_element = browser_locate_element_s(contact_xpath)
        browser_mouse_click_s(contact_element)
        browser_wait_s()
        message_element = browser_locate_element_s(message_xpath)
        browser_wait_s()
        time.sleep(5)
        browser_write_s(message_element,message)
        browser_mouse_click_s(message_element,i="e")
        time.sleep(1)
        browser_wait_s()
        driver.close()

    except Exception as ex:
            print("Error in message_send_whatsapp()="+str(ex))

def excel_draw_charts(excel_path="",sheet_name='Sheet1', header=0, x_col="", y_col="", color="", chart_type='bar', title='ClointFusion', show_chart=False):

    """
    Interactive data visualization function, which accepts excel file, X & Y column. 
    Chart types accepted are bar , scatter , pie , sun , histogram , box  , strip. 
    You can pass color column as well, having a boolean value.
    Image gets saved as .PNG in the same path as excel file.

    Usage: excel_charts(<excel path>,x_col='Name',y_col='Age', chart_type='bar',show_chart=True)
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('for data visualization')
            
        if not x_col:
            # x_col = gui_get_any_input_from_user("X Axis Column")
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            x_col = gui_get_dropdownlist_values_from_user('X Axis Column',col_lst)  

        if not y_col:
            # y_col = gui_get_any_input_from_user("Y Axis Column")
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            y_col = gui_get_dropdownlist_values_from_user('Y Axis Column',col_lst)  

        if x_col and y_col:
            if color:
                df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,usecols={x_col,y_col,color})
            else:
                df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,usecols={x_col,y_col})

            fig = go.Figure()

            if chart_type == 'bar':

                fig.add_trace(go.Bar(x=df[x_col].values.tolist()))
                fig.add_trace(go.Bar(y=df[y_col].values.tolist()))

                if color:
                    fig = px.bar(df, x=x_col, y=y_col, barmode="group",color=color)
                else:
                    fig = px.bar(df, x=x_col, y=y_col, barmode="group")
                    
            elif chart_type == 'scatter':

                fig.add_trace(go.Scatter(x=df[x_col].values.tolist()))
                fig.add_trace(go.Scatter(y=df[x_col].values.tolist()))

            elif chart_type =='pie':

                if color:
                    fig = px.pie(df, names=x_col, values=y_col, title=title,color=color)#,hover_data=df.columns)
                else:
                    fig = px.pie(df, names=x_col, values=y_col, title=title)#,hover_data=df.columns)

            elif chart_type =='sun':

                if color:
                    fig = px.sunburst(df, path=[x_col], values=y_col,hover_data=df.columns,color=color)
                else:
                    fig = px.sunburst(df, path=[x_col], values=y_col,hover_data=df.columns)

            elif chart_type == 'histogram':

                if color:
                    fig = px.histogram(df, x=x_col, y=y_col, marginal="rug",color=color, hover_data=df.columns)
                else:
                    fig = px.histogram(df, x=x_col, y=y_col, marginal="rug",hover_data=df.columns)

            elif chart_type == 'box':

                if color:
                    fig = px.box(df, x=x_col, y=y_col, notched=True,color=color)
                else:
                    fig = px.box(df, x=x_col, y=y_col, notched=True)

            elif chart_type == 'strip':

                if color:
                    fig = px.strip(df, x=x_col, y=y_col, orientation="h",color=color)
                else:
                    fig = px.strip(df, x=x_col, y=y_col, orientation="h")

            fig.update_layout(title = title)
            
            if show_chart:
                fig.show()
            
            # strFileName = (excel_path[excel_path.rindex("\\")+1:]).split(".")[0] + ".PNG"
            strFileName = excel_path.replace(".xlsx",".PNG")
            
            scope = PlotlyScope()
            with open(strFileName, "wb") as f:
                f.write(scope.transform(fig, format="png"))
            print("Chart saved at " + strFileName)
        else:
            print("Please supply all the required values")

    except Exception as ex:
        print("Error in excel_draw_charts=" + str(ex))

def get_long_lat(strZipCode=0):
    """
    Function takes zip_code as input (int) and returns longitude, latitude, state, city, county. 
    """
    try:
        if not strZipCode:
            strZipCode = str(gui_get_any_input_from_user("USA Zip Code ex: 77429"))

        all_data_dict=zipcodes.matching(str(strZipCode))

        all_data_dict = all_data_dict[0]

        long = all_data_dict['long']
        lat = all_data_dict['lat']
        state = all_data_dict['state']
        city = all_data_dict['city']
        county = all_data_dict['county']
        return long, lat, state, city, county    
    except Exception as ex:
        print("Error in get_long_lat="+str(ex))

def excel_geotag_using_zipcodes(excel_path="",sheet_name='Sheet1',header=0,zoom_start=5,zip_code_column="ZIP CODE",data_columns_as_list=[],color_boolean_column=""):
    """
    Function takes Excel file having ZipCode column as input. Takes one data column at present. 
    Creates .html file having geo-tagged markers/baloons on the page.

    Ex: excel_geotag_using_zipcodes(excel_path)
    """

    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('for geo tagging')

        m = folium.Map(location=[40.178877,-100.914253 ], zoom_start=zoom_start)

        if len(data_columns_as_list) == 1:
            data_columns_as_str = str(data_columns_as_list).replace("[","").replace("]","").replace("'","")
        else:
            data_columns_as_str = str(data_columns_as_list).replace("[","").replace("]","")
            data_columns_as_str = data_columns_as_str[1:-1]
            
        use_cols = data_columns_as_list
        use_cols.append(zip_code_column)

        if color_boolean_column:
            use_cols.append(color_boolean_column)

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,usecols=use_cols)
        
        for _, row in df.iterrows():
            if not pd.isna(row[zip_code_column]) and str(row[zip_code_column]).isnumeric():
                
                long, lat, state, city, county = get_long_lat(str(row[zip_code_column]))
                county = str(county).replace("County","")
                
                if color_boolean_column and data_columns_as_str and row[color_boolean_column] == True:
                    folium.Marker(location=[lat, long], popup='State: ' + state + ',\nCity:' + city + ',\nCounty:' + county + ',\nDevice:' + row[data_columns_as_str], icon=folium.Icon(color='green', icon='info-sign')).add_to(m)
                elif data_columns_as_str:
                    folium.Marker(location=[lat, long], popup='State: ' + state + ',\nCity:' + city + ',\nCounty:' + county + ',\nDevice:' + row[data_columns_as_str], icon=folium.Icon(color='red', icon='info-sign')).add_to(m)
                else:
                    folium.Marker(location=[lat, long], popup='State: ' + state + ',\nCity:' + city + ',\nCounty:' + county, icon=folium.Icon(color='blue', icon='info-sign')).add_to(m)

        graphFileName = excel_path.replace(".xlsx",".html")
        print("GeoTagged Graph saved at "+ graphFileName)
        m.save(graphFileName)
    
    except Exception as ex:
        print("Error in excel_geotag_using_zipcodes="+str(ex))
    
def _accept_cookies_h():
    """
    Internal function to accept cookies.
    """
    try:
        if Text('Accept cookies?').exists():
            click('I accept')
    except Exception as ex:
        print("Error in _accept_cookies_h="+str(ex))
    
def launch_website_h(URL="",dp=False,dn=True,igc=True,smcp=True,i=False,headless=False):
    try:
        """
        Internal function to launch browser.
        """
        if not URL:
            URL = gui_get_any_input_from_user("website URL to Launch Website using Helium functions. Ex https://www.google.com")
        global HLaunched
        HLaunched=True
        options = ChromeOptions()
        if dp:
            options.add_argument("--disable-popup-blocking")                
        if dn:  
            options.add_argument("--disable-notifications")                
        if igc:
            options.add_argument("--ignore-certificate-errors")             
        if smcp:
            options.add_argument("--suppress-message-center-popups")       
        if i:
            options.add_argument("--incognito")                             
        
        options.add_argument("--disable-translate")
        options.add_argument("--start-maximized")                          
        options.add_argument("--ignore-autocomplete-off-autofill")          
        options.add_argument("--no-first-run")                             
        #options.add_argument("--window-size=1920,1080")
        start_chrome(url=URL,options=options,headless=headless)
        Config.implicit_wait_secs = 120
        _accept_cookies_h()
    except Exception as ex:
        print("Error in launch_website_h = "+str(ex))
        kill_browser()
    
def browser_navigate_h(url="",dp=False,dn=True,igc=True,smcp=True,i=False,headless=False):
    try:
        """
        Navigates to Specified URL.
        """
        if not url:
            url = gui_get_any_file_from_user("website URL to Navigate using Helium functions. Ex: https://www.google.com")

        global HLaunched
        if not HLaunched:
            launch_website_h(URL=url,dp=dp,dn=dn,igc=igc,smcp=smcp,i=i,headless=headless)
            return
        go_to(url.lower())
        _accept_cookies_h()
    except Exception as ex:
        print("Error in browser_navigate_h = "+str(ex))
    
def browser_write_h(Value="",User_Visible_Text_Element="",alert=False):
    """
    Write a string on the given element.
    """
    try:
        if not User_Visible_Text_Element:
            User_Visible_Text_Element = gui_get_any_input_from_user('visible element (placeholder) to WRITE your value. Ex: Username')

        if not Value:
            Value= gui_get_any_input_from_user('Value to be Written')

        if not alert:
            if Value and User_Visible_Text_Element:
                write(Value, into=User_Visible_Text_Element)
        if alert:
            if Value and User_Visible_Text_Element:
                write(Value, into=Alert(User_Visible_Text_Element))
    except Exception as ex:
        print("Error in browser_write_h = "+str(ex))
    
# def browser_key_press_h(key=""):
#     """
#     keyboard simulation.
#     """
#     try:
#         if not key:
#             key = gui_get_any_input_from_user('key to be pressed')

#         press(key)
#     except Exception as ex:
#         print("Error in browser_key_press_h="+str(ex))
    
def browser_mouse_click_h(User_Visible_Text_Element="",element="d"):
    """
    click on the given element.
    """
    try:
        if not User_Visible_Text_Element:
            User_Visible_Text_Element = gui_get_any_input_from_user("visible text element (button/link/checkbox/radio etc) to Click")

        if User_Visible_Text_Element and element.lower()=="d":      #default
            click(User_Visible_Text_Element)
        elif User_Visible_Text_Element and element.lower()=="l":    #link
            click(link(User_Visible_Text_Element))
        elif User_Visible_Text_Element and element.lower()=="b":    #button
            click(Button(User_Visible_Text_Element))
        elif User_Visible_Text_Element and element.lower()=="t":    #textfeild
            click(TextField(User_Visible_Text_Element))
        elif User_Visible_Text_Element and element.lower()=="c":    #checkbox
            click(CheckBox(User_Visible_Text_Element))
        elif User_Visible_Text_Element and element.lower()=="r":    #radiobutton
            click(RadioButton(User_Visible_Text_Element))
        elif User_Visible_Text_Element and element.lower()=="i":    #image ALT Text
            click(Image(alt=User_Visible_Text_Element))
    except Exception as ex:
        print("Error in browser_mouse_click_h = "+str(ex))
    

def browser_mouse_double_click_h(User_Visible_Text_Element=""):
    """
    Doubleclick on the given element.
    """
    try:
        if not User_Visible_Text_Element:
            User_Visible_Text_Element = gui_get_any_input_from_user("visible text element (button/link/checkbox/radio etc) to Double Click")

        if User_Visible_Text_Element:
            doubleclick(User_Visible_Text_Element)
    except Exception as ex:
        print("Error in browser_mouse_double_click_h = "+str(ex))
    
        
def browser_locate_element_h(element="",value=True):
    """
    Find the element by Xpath, id or css selection.
    """
    try:
        if not element:
            element = gui_get_any_input_from_user('browser element to locate (Helium)')
        if value:
            return S(element).value
        return S(element)
    except Exception as ex:
        print("Error in browser_locate_element_h = "+str(ex))
    
def browser_locate_elements_h(element="",value=True):
    """
    Find the elements by Xpath, id or css selection.
    """
    try:
        if not element:
            element = gui_get_any_input_from_user('browser ElementS to locate (Helium)')
        if value:
            return find_all(S(element).value)
        return find_all(S(element))
    except Exception as ex:
        print("Error in browser_locate_elements_h = "+str(ex))
    
def browser_wait_until_h(text="",element="t"):
    """
    Wait until a specific element is found.
    """
    try:
        if not text:
            text = gui_get_any_input_from_user("visible text element to Search & Wait for")

        if element.lower()=="t":
            wait_until(Text(text).exists,10)        #text
        elif element.lower()=="b":
            wait_until(Button(text).exists,10)      #button
    except Exception as ex:
        print("Error in browser_wait_until_h = "+str(ex))

def browser_mouse_click_xy_h(XYTuple=""):
    """
    Click on the given X Y Co-ordinates.
    """
    try:
        if not XYTuple:
            XYTuple1 = gui_get_any_input_from_user('browser X,Y co-ordinates for Mouse Left Click. Ex: (300,400)')
            XYTuple = tuple(map(int, XYTuple1.split(',')))
        
        click(XYTuple)
    except Exception as ex:
        print("Error in browser_mouse_click_xy_h = "+str(ex))
    
def browser_refresh_page_h():
    """
    Refresh the page.
    """
    try:
        refresh()
    except Exception as ex:
        print("Error in browser_refresh_page_h = "+str(ex))
    
def browser_quit_h():
    """
    Close the browser.
    """
    try:
        kill_browser()
    except Exception as ex:
        print("Error in browser_quit_h = "+str(ex))
    

#Utility Functions
def dismantle_code(strFunctionName=""):
    """
    This functions dis-assembles given function and shows you column-by-column summary to explain the output of disassembled bytecode.
    Ex: dismantle_code(show_emoji)
    """
    try:
        if not strFunctionName:
            strFunctionName = gui_get_any_input_from_user('function name to dis-assemble')
            print("Code dismantling {}".format(strFunctionName))
            return dis.dis(strFunctionName) 
    except Exception as ex:
       print("Error in dismantle_code="+str(ex)) 

def excel_clean_data(excel_path="",sheet_name='Sheet1',header=0,column_to_be_cleaned="",cleaning_pipe_line="Default"):
    """
    fillna(s) Replace not assigned values with empty spaces.
    lowercase(s) Lowercase all text.
    remove_digits() Remove all blocks of digits.
    remove_punctuation() Remove all string.punctuation (!"#$%&'()*+,-./:;<=>?@[\]^_`{|}~).
    remove_diacritics() Remove all accents from strings.
    remove_stopwords() Remove all stop words.
    remove_whitespace() Remove all white space between words.
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to clean the data')
            
        if not column_to_be_cleaned:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)  
            column_to_be_cleaned = gui_get_dropdownlist_values_from_user('column list to Clean',col_lst,multi_select=False)   
            column_to_be_cleaned = column_to_be_cleaned[0]

        if column_to_be_cleaned:
            df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header)

            new_column_name = "Clean_" + column_to_be_cleaned

            if 'Default' in cleaning_pipe_line:
                df[new_column_name] = df[column_to_be_cleaned].pipe(hero.clean)
            else:
                custom_pipeline = [preprocessing.fillna, preprocessing.lowercase]
                df[new_column_name] = df[column_to_be_cleaned].pipe(hero.clean,custom_pipeline)    

            with pd.ExcelWriter(path=excel_path, mode='a',engine='openpyxl') as writer:
                df.to_excel(writer,index=False)

            print("Data Cleaned. Please see the output in {}".format(new_column_name))
    except Exception as ex:
        print("Error in excel_clean_data="+str(ex))
    
def excel_charts_numerical_pair_plot(excel_path="", sheet_name="", header=0, usecols=""):
    """
    Funtion for statistical data visualization.
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('for pair plot charts')

        if not usecols:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            usecols = gui_get_dropdownlist_values_from_user('column list, to used for Pair Plot',col_lst)   

        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols=usecols)
        
        # df = sb.load_dataset('iris')
        g = sb.PairGrid(df)
        g.map(plt.scatter)
        g.map_diag(plt.hist)
        g.map_offdiag(plt.scatter)
        
        mng = plt.get_current_fig_manager()
        mng.full_screen_toggle()
        plt.tight_layout()
        plt.show()

        strFileName = excel_path.replace(".xlsx",".PNG")
        plt.savefig(strFileName)
    
        print("Chart saved at " + strFileName)

    except Exception as ex:
        print("Error in excel_charts_numerical_pair_plot="+str(ex))

def excel_charts_correlation_heatmap(excel_path="", sheet_name="", header=0, usecols=""):
    """
    Function for co-relation between columns of the given excel
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('for correlation heatmap charts')

        if not usecols:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            usecols = gui_get_dropdownlist_values_from_user('column list to used for Co-relation heatmap',col_lst)  

        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols=usecols)

        # print(df.corr())
        sb.heatmap(df.corr(), annot = True, cmap = 'viridis')
        
        strFileName = excel_path.replace(".xlsx",".PNG")
        
        plt.savefig(strFileName)
        print("Chart saved at " + strFileName)

        mng = plt.get_current_fig_manager()
        mng.full_screen_toggle()
        plt.tight_layout()
        plt.show()
    except Exception as ex:
        print("Error in excel_charts_correlation_heatmap="+str(ex))
    
def help_me(user_query=""):
    """
    Function which gives instant coding answers

    Usage:

    help_me("convert mp4 to animated gif")
    help_me("create tar archive")
    help_me("howdoi")
    help_me("formatting date")
    """
    try:
        if not user_query:
            user_query = gui_get_any_input_from_user("your Query. Ex: formatting date")

        print("Output from instant coding answers:")
        os.system("howdoi {}".format(user_query))
        print(show_emoji('point_up'))
    except Exception as ex:
        print("Error in help_me="+str(ex))

def browser_get_header_source_code(URL=""):
    """
    Function to get Header and Source Code for the given URL
    """
    try:
        if not URL:
            URL = gui_get_any_input_from_user("website URL to get Header & Source Code. Ex: https://www.cloint.com")

        page = urlopen(URL) 
        print("\nPage Headers\n")
        print(page.headers) 
        
        content = page.read() 

        file_name = string_remove_special_characters(URL) + ".html"
        file_path = os.path.join(output_folder_path,file_name)

        with open(file_path,"wb") as f:
            f.write(content)
            
        print("Source code saved at "+ str(file_path))
        return file_path
    except Exception as ex:
        print("Error in browser_get_header_source_code"+str(ex))

def compute_hash(inputData=""):
    """
    Returns the hash of the inputData 
    """
    try:
        if not inputData:
            inputData = gui_get_any_input_from_user('input string to compute Hash')

        return sha256(inputData.encode()).hexdigest()
    except Exception as ex:
        print("Error in compute_hash="+str(ex))

def browser_get_html_text(url=""):
    """
    Function to get HTML text without tags using Beautiful soup
    """
    try:
        if not url:
            url = gui_get_any_input_from_user("website URL to get HTML Text (without tags). Ex: https://www.cloint.com")

        html_text = requests.get(url) 
        soup = BeautifulSoup(html_text.content, 'lxml')
        text = str(soup.text).strip()
        text = ' '.join(text.split())
        return text
    except Exception as ex:
        print("Error in browser_get_html_text="+str(ex))

def word_cloud_from_url(url=""):
    """
    Function to create word cloud from a given website
    """
    try:
        text = browser_get_html_text(url=url)
        
        wc = WordCloud(max_words=2000, width=800, height=600,background_color='white',max_font_size=40, random_state=None, relative_scaling=0)
        wc.generate(text)
        file_path = os.path.join(output_folder_path,"URL_WordCloud.png")
        wc.to_file(file_path)
        print("URL WordCloud saved at {}".format(file_path))

    except Exception as ex:
        print("Error in word_cloud_from_url="+str(ex))

def word_cloud_from_excel(excel_path="",sheet_name="",header=0,columnName=""):
    """
    Function to create word cloud from a given website
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('for WordCloud')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)

            columnName = gui_get_dropdownlist_values_from_user('list of Column names',col_lst)    

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header)

        text = ""

        if columnName:
            text = ''.join(str(df[columnName]))
            text = text.replace("\n"," ")
        
        wc = WordCloud(max_words=2000, width=800, height=600, max_font_size=40, random_state=None, relative_scaling=0)
        wc.generate(text)
        file_path = os.path.join(output_folder_path,"Excel_WordCloud.png")
        wc.to_file(file_path)
        print("Excel WordCloud saved at {}".format(file_path))

    except Exception as ex:
        print("Error in word_cloud_from_excel="+str(ex))

def pdf_text_extract(path_to_pdf=""):
    """
    Extract data from PDF(s). Works best on machine-generated PDF, than scanned.
    """
    try:
        if not path_to_pdf:
            path_to_pdf = gui_get_any_file_from_user("the PDF to extract text","pdf")

        pdf_text = ""

        with pdfplumber.open(path_to_pdf) as pdf:
            total_pages = int(len(pdf.pages))
            print("Total PDF page(s)="+str(total_pages))
            for i in range(total_pages):
                page = pdf.pages[i]
                pdf_text = pdf_text + page.extract_text()
                
        return pdf_text

    except Exception as ex:
        print("Error in pdf_text_extract="+str(ex))

def word_cloud_from_pdf(path_to_pdf=""):
    """
    Function to create word cloud from a given PDF
    """
    try:

        text = pdf_text_extract(path_to_pdf=path_to_pdf)
        
        wc = WordCloud(max_words=2000, width=800, height=600, max_font_size=40, random_state=None, background_color='black',relative_scaling=0)
        wc.generate(text)
        file_path = os.path.join(output_folder_path,"PDF_WordCloud.png")
        wc.to_file(file_path)
        print("PDF WordCloud saved at {}".format(file_path))

    except Exception as ex:
        print("Error in word_cloud_from_pdf="+str(ex))

def excel_describe_data(excel_path="",sheet_name='Sheet1',header=0):
    """
    Describe statistical data for the given excel
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user("to Statistically Describe excel data")
            
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header)

        user_option_lst = ['Numerical','String','Both']

        user_choice = gui_get_dropdownlist_values_from_user("list of datatypes",user_option_lst)

        if user_choice == 'Numerical':
            return df.describe(include = [np.number])
        elif user_choice == 'String':
            return df.describe(include = ['O'])
        else:
            return df.describe(include='all')

    except Exception as ex:
        print("Error in excel_describe_data="+str(ex))

def camera_capture_image(user_name=""):
    try:

        user_consent = gui_get_consent_from_user("turn ON camera & take photo ?")

        if user_consent == 'Yes':
            SECONDS = 5
            TIMER = int(SECONDS) 
            window_name = "ClointFusion"
            cap = cv2.VideoCapture(0) 

            if not cap.isOpened():
                print("Error in opening camera")

            cv2.namedWindow(window_name, cv2.WND_PROP_FULLSCREEN)
            cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
            font = cv2.FONT_HERSHEY_SIMPLEX 

            if not user_name:
                user_name = gui_get_any_input_from_user("your name")

            while True: 

                ret, img = cap.read() 
                cv2.imshow(window_name, img) 
                prev = time.time() 

                text = "Taking selfie in 5 second(s)".format(str(TIMER))
                textsize = cv2.getTextSize(text, font, 1, 2)[0]
                print(str(textsize))

                textX = int((img.shape[1] - textsize[0]) / 2)
                textY = int((img.shape[0] + textsize[1]) / 2)

                while TIMER >= 0: 
                    ret, img = cap.read() 

                    cv2.putText(img, "Saving image in {} second(s)".format(str(TIMER)),  
                                (textX, textY ), font, 
                                1, (255, 0, 255), 
                                2) 
                    cv2.imshow(window_name, img) 
                    cv2.waitKey(125) 

                    cur = time.time() 

                    if cur-prev >= 1: 
                        prev = cur 
                        TIMER = TIMER-1

                ret, img = cap.read() 
                cv2.imshow(window_name, img) 
                cv2.waitKey(1000) 
                file_path = os.path.join(output_folder_path,user_name + ".PNG")
                cv2.imwrite(file_path, img) 
                print("Image saved at {}".format(file_path))
                cap.release() 
                cv2.destroyAllWindows()
                break

        else:
            print("Operation cancelled by user")

    except Exception as ex:
        print("Error in camera_capture_image="+str(ex))   

def excel_set_formula(excel_path="",sheet_name='Sheet1',target_cell="",formula="",cell_format='General'):
    """
    Function to set formula to a particular excel cell
    """
    try:
        if not excel_path:
            excel_path, sheet_name, _ = gui_get_excel_sheet_header_from_user('to set Formula')

        if not target_cell:
            target_cell = gui_get_any_input_from_user("target Cell Ex: B6")

        if not formula:
            formula = gui_get_any_input_from_user("formula Ex: =SUM(B2:B5)")

            wb = op.load_workbook(excel_path)
            ws = wb[sheet_name]

            ws[target_cell] = formula

            ws[target_cell].number_format = cell_format
            
            wb.save(excel_path)
    except Exception as ex:
        print("Error in excel_set_formula="+str(ex))             

def convert_csv_to_excel(csv_path="",sep=""):
    """
    Function to convert CSV to Excel 

    Ex: convert_csv_to_excel()
    """
    try:
        if not csv_path:
            csv_path = gui_get_any_file_from_user("CSV to convert to EXCEL","csv")

        if not sep:
            sep = gui_get_any_input_from_user("Delimeter Ex: |")

        csv_file_name = extract_filename_from_filepath(csv_path)
        excel_file_name = csv_file_name + ".xlsx"        
        df=pd.read_csv(csv_path,sep=sep)

        excel_file_path = os.path.join(output_folder_path,excel_file_name)
        writer = pd.ExcelWriter(excel_file_path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        writer.save()
        writer.close()
        print("Excel file saved : "+str(excel_file_path))

    except Exception as ex:
        print("Error in convert_csv_to_excel="+str(ex))

def email_send_outlook_desktop(TO="",CC="",BCC="",SUBJECT="",BODY="",ATTACHMENT_FILE_PATH=""):
    """
    Function to send email via Outlook Desktop App. This uses already logged in Outlook Desktop App.

    Ex: email_send_outlook_desktop()
    """
    try:
        if not TO:
            TO = gui_get_any_input_from_user("Comma separated `To` email IDs")
        if not CC:
            CC = gui_get_any_input_from_user("Comma separated `CC` email IDs",mandatory_field=False)
        if not BCC:
            BCC = gui_get_any_input_from_user("Comma separated `BCC` email IDs",mandatory_field=False)
        if not SUBJECT:
            SUBJECT = gui_get_any_input_from_user("Email Subject")
        if not BODY:
            BODY = gui_get_any_input_from_user("Email Body")
        if not ATTACHMENT_FILE_PATH:
            ATTACHMENT_FILE_PATH = gui_get_any_file_from_user("Attachment File")

        try:
            launch_any_exe_bat_application(r"C:\Program Files (x86)\Microsoft Office\root\Office16\OUTLOOK.EXE")
            time.sleep(5)
        except:
            key_hit_enter()
            launch_any_exe_bat_application(r"C:\Program Files (x86)\Microsoft Office\root\Office16\OUTLOOK.EXE")
            time.sleep(5)

        key_press('ctrl+shift+M') #compose new email
        time.sleep(2)
        
        key_write_enter(TO ,delay=1, key="t")
        key_press("tab")
        
        if CC:
            key_write_enter(CC ,delay=1, key="t")

        key_press("tab")
        
        if BCC:
            key_write_enter(BCC ,delay=2, key="t")
            key_press("tab")
        
        key_write_enter(SUBJECT,delay=2, key="t")
        
        key_write_enter("Hi There!",delay=1)
        key_hit_enter()

        key_write_enter(BODY,delay=2)

        key_hit_enter()

        key_write_enter("Best Regards,",delay=2)

        key_press("alt+h")
        key_press("a+f")
        key_press("b")
        time.sleep(2)
        ATTACHMENT_FILE_PATH = ATTACHMENT_FILE_PATH.replace("/","\\")
        key_write_enter(ATTACHMENT_FILE_PATH)
        key_press("ctrl+enter")
        time.sleep(5)
        key_press("alt+f4")
        time.sleep(2)
        
    except Exception as ex:
        print("Error in send_email_outlook" + str(ex))

def watch_this_folder(folder_to_watch=""):
    """
    Function to Monitor the given folder for creation / modification / deletion events. You can take required action, as per usecase
    Ex: watch_this_folder()
    """
    try:
        if not folder_to_watch:
            folder_to_watch= gui_get_folder_path_from_user('folder to Watch / Monitor')

        event_handler = FileMonitor_Handler()
        observer = watchdog.observers.Observer()
        observer.schedule(event_handler,folder_to_watch, recursive = False)
        observer.start()

        try:
            print("Monitoring Folder: {} every 1 Second, for create / modify / delete events".format(folder_to_watch))
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("Stopping Folder Monitor...")
            observer.stop()

        observer.join()
        
    except Exception as ex:
        print("Error in watch_this_folder")

# Class related to capture_snip_now
class CaptureSnip(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        root = tk.Tk()
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        self.setGeometry(0, 0, screen_width, screen_height)
        self.setWindowTitle(' ')
        self.begin = QtCore.QPoint()
        self.end = QtCore.QPoint()
        self.setWindowOpacity(0.3)
        QtWidgets.QApplication.setOverrideCursor(
            QtGui.QCursor(QtCore.Qt.CrossCursor)
        )
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        print('Capture now...')
        self.show()

    def paintEvent(self, event):
        qp = QtGui.QPainter(self)
        qp.setPen(QtGui.QPen(QtGui.QColor('black'), 3))
        qp.setBrush(QtGui.QColor(128, 128, 255, 128))
        qp.drawRect(QtCore.QRect(self.begin, self.end))

    def mousePressEvent(self, event):
        self.begin = event.pos()
        self.end = self.begin
        self.update()

    def mouseMoveEvent(self, event):
        self.end = event.pos()
        self.update()

    def mouseReleaseEvent(self, event):
        self.close()

        x1 = min(self.begin.x(), self.end.x())
        y1 = min(self.begin.y(), self.end.y())
        x2 = max(self.begin.x(), self.end.x())
        y2 = max(self.begin.y(), self.end.y())

        img = ImageGrab.grab(bbox=(x1, y1, x2, y2))
        file_num = str(len(os.listdir(img_folder_path)))
        file_name = os.path.join(img_folder_path,file_num + "_snip.PNG" )
        img.save(file_name)
        return file_name
        
def capture_snip_now():
    """
    Captures the snip and stores in Image Folder of the BOT by giving continous numbering

    Ex: capture_snip_now()
    """
    try:
        if message_counter_down_timer(3):
            app = QtWidgets.QApplication(sys.argv)
            window = CaptureSnip()
            window.activateWindow()
            app.aboutToQuit.connect(app.deleteLater)
            sys.exit(app.exec_())
            
    except Exception as ex:
        print("Error in capture_snip_now="+str(ex))        

def ON_semi_automatic_mode():
    """
    This function sets semi_automatic_mode as True => ON
    """
    global enable_semi_automatic_mode
    semi_automatic_config_file_path = os.path.join(config_folder_path,"Semi_Automatic_Mode.txt")
    try:    
        with open(semi_automatic_config_file_path, 'w') as f:
            f.write('True')
        enable_semi_automatic_mode = True
        print("Semi Automatic Mode is ENABLED "+ show_emoji())
    except Exception as ex:
        print("Error in ON_semi_automatic_mode="+str(ex))

def OFF_semi_automatic_mode():
    """
    This function sets semi_automatic_mode as False => OFF
    """
    global enable_semi_automatic_mode
    semi_automatic_config_file_path = os.path.join(config_folder_path,"Semi_Automatic_Mode.txt")
    try:    
        with open(semi_automatic_config_file_path, 'w') as f:
            f.write('False')
        enable_semi_automatic_mode = False
        print("Semi Automatic Mode is DISABLED "+ show_emoji())
    except Exception as ex:
        print("Error in OFF_semi_automatic_mode="+str(ex))        